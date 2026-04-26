# -*- coding: utf-8 -*-
"""
问题1：混合车队日常配送调度启发式求解脚本

输入文件（需与本脚本放在同一目录）：
1. 订单信息_补全版.xlsx
2. 距离矩阵.xlsx
3. 客户坐标信息.xlsx
4. 时间窗.xlsx

输出文件：
- q1_result.xlsx：汇总表、线路表、停靠表、车辆复用表、批次表等
- q1_transport_flow_map.png：运输流程/路线图
- q1_vehicle_gantt.png：车辆运行甘特图
- q1_cost_breakdown.png：成本结构图
- q1_console_summary.txt：控制台汇总结果

模型要点：
- 客户需求先按 3000kg 与燃油车 13.5m^3 容量拆分为可配送批次；拆分时优先形成满载批次，最后保留剩余批次参与拼车；
- 每条线路最多 5 个批次停靠点；
- 行驶速度按题设时段分段，计算时间依赖行驶时间；
- 等待成本 20 元/小时，迟到成本 50 元/小时；
- 燃油/电动车能源成本与碳排成本按速度函数和载重率修正；
- 先构造线路，再用容量为 10 的加权区间选择分配新能源车，其余分配燃油车；
- 最后按线路时间区间复用实体车辆，固定成本按实体车辆首次启用计入。
"""

from __future__ import annotations

import heapq
import itertools
import math
import sys
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# =========================
# 1. 文件与参数
# =========================
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR

ORDER_FILE = "订单信息_补全版.xlsx"
DIST_FILE = "距离矩阵.xlsx"
COORD_FILE = "客户坐标信息.xlsx"
TW_FILE = "时间窗.xlsx"

DEPOT_ID = 0
START_OF_DAY = 8 * 60
END_OF_DAY = 24 * 60
SERVICE_MIN = 20
MAX_ROUTE_STOPS = 5

# 3000kg 车型参数
VEHICLE_WEIGHT_CAP = 3000.0
FUEL_VOLUME_CAP = 13.5
EV_VOLUME_CAP = 15.0
EV_FLEET_LIMIT = 10
FUEL_FLEET_LIMIT = 60
FIXED_VEHICLE_COST = 400.0

# 价格、碳排系数与时间窗成本
FUEL_PRICE = 7.61              # 元/L
ELECTRIC_PRICE = 1.64          # 元/kWh
CARBON_PRICE = 0.65            # 元/kgCO2
FUEL_CARBON_FACTOR = 2.547     # kgCO2/L
ELECTRIC_CARBON_FACTOR = 0.501 # kgCO2/kWh
WAIT_COST_PER_HOUR = 20.0
LATE_COST_PER_HOUR = 50.0

# 用于构造线路时的近似固定成本。实际结算按实体车辆首次启用计。
APPROX_FIXED_COST_PER_ROUTE = 400.0
CROSS_DAY_PENALTY_PER_MIN = 5000.0  # 不跨天硬约束的惩罚系数

# 时间依赖速度：单位 min, km/h
TRAFFIC_INTERVALS = [
    (0, 8 * 60, 35.4),
    (8 * 60, 9 * 60, 9.8),
    (9 * 60, 10 * 60, 55.3),
    (10 * 60, int(11.5 * 60), 35.4),
    (int(11.5 * 60), 13 * 60, 9.8),
    (13 * 60, 15 * 60, 55.3),
    (15 * 60, 24 * 60, 35.4),
    (24 * 60, 10**9, 35.4),
]

# 结果文件名
RESULT_XLSX = OUTPUT_DIR / "q1_result_fullsplit.xlsx"
FLOW_PNG = OUTPUT_DIR / "q1_transport_flow_map_fullsplit.png"
GANTT_PNG = OUTPUT_DIR / "q1_vehicle_gantt_fullsplit.png"
COST_PNG = OUTPUT_DIR / "q1_cost_breakdown_fullsplit.png"
SUMMARY_TXT = OUTPUT_DIR / "q1_console_summary_fullsplit.txt"


# =========================
# 2. 通用工具函数
# =========================
def hhmm_to_minute(x: Any) -> int:
    """将 HH:MM、datetime.time 或 Excel 小数时间转为当天分钟数。"""
    if pd.isna(x):
        raise ValueError("时间窗存在空值。")
    if hasattr(x, "hour") and hasattr(x, "minute"):
        return int(x.hour) * 60 + int(x.minute)
    if isinstance(x, (int, float, np.integer, np.floating)):
        # Excel 日期小数：一天=1。若已是分钟，直接返回。
        val = float(x)
        if 0 <= val <= 1:
            return int(round(val * 24 * 60))
        return int(round(val))
    s = str(x).strip()
    if " " in s:
        s = s.split()[-1]
    parts = s.split(":")
    if len(parts) < 2:
        raise ValueError(f"无法识别时间格式: {x}")
    return int(parts[0]) * 60 + int(parts[1])


def minute_to_hhmm(x: float) -> str:
    """将分钟数转为 HH:MM，可超过 24:00。"""
    x = float(x)
    if not np.isfinite(x):
        return ""
    h = int(x // 60)
    m = int(round(x % 60))
    if m == 60:
        h += 1
        m = 0
    return f"{h:02d}:{m:02d}"


def safe_round(x: Any, ndigits: int = 4) -> Any:
    try:
        if pd.isna(x):
            return x
        return round(float(x), ndigits)
    except Exception:
        return x


def fuel_per_100km(v: float) -> float:
    """燃油车百公里油耗函数，单位 L/100km。"""
    return 0.0025 * v * v - 0.2554 * v + 31.75


def electric_per_100km(v: float) -> float:
    """新能源车百公里电耗函数，单位 kWh/100km。"""
    return 0.0014 * v * v - 0.12 * v + 36.19


def set_matplotlib_font() -> None:
    """尽量避免中文缺字。若环境没有中文字体，图中使用英文标签。"""
    plt.rcParams["font.sans-serif"] = ["DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["figure.dpi"] = 140
    plt.rcParams["savefig.dpi"] = 220


# =========================
# 3. 数据读取与批次拆分
# =========================
def complete_order_weight_volume(orders: pd.DataFrame) -> pd.DataFrame:
    """根据重量和体积的经验比例补全订单缺失值。

    对重量和体积均有效的订单，计算总体积/总重量比例 lambda。
    - 仅体积缺失：volume = lambda * weight
    - 仅重量缺失：weight = volume / lambda
    两项均缺失或非正的记录不参与有效需求统计。
    """
    orders = orders.copy()
    orders["weight_kg"] = pd.to_numeric(orders["weight_kg"], errors="coerce")
    orders["volume_m3"] = pd.to_numeric(orders["volume_m3"], errors="coerce")

    valid_weight = orders["weight_kg"].notna() & (orders["weight_kg"] > 0)
    valid_volume = orders["volume_m3"].notna() & (orders["volume_m3"] > 0)
    complete = valid_weight & valid_volume
    if complete.sum() == 0:
        raise ValueError("订单数据中没有可用于估计重量-体积关系的完整记录。")

    lambda_v_per_kg = float(orders.loc[complete, "volume_m3"].sum() / orders.loc[complete, "weight_kg"].sum())
    if lambda_v_per_kg <= 0 or not np.isfinite(lambda_v_per_kg):
        raise ValueError("重量-体积换算比例异常，无法补全缺失数据。")

    orders["weight_imputed"] = False
    orders["volume_imputed"] = False

    missing_volume = valid_weight & ~valid_volume
    orders.loc[missing_volume, "volume_m3"] = orders.loc[missing_volume, "weight_kg"] * lambda_v_per_kg
    orders.loc[missing_volume, "volume_imputed"] = True

    valid_volume_after = orders["volume_m3"].notna() & (orders["volume_m3"] > 0)
    missing_weight = ~valid_weight & valid_volume_after
    orders.loc[missing_weight, "weight_kg"] = orders.loc[missing_weight, "volume_m3"] / lambda_v_per_kg
    orders.loc[missing_weight, "weight_imputed"] = True

    orders["weight_kg"] = orders["weight_kg"].fillna(0.0).clip(lower=0.0)
    orders["volume_m3"] = orders["volume_m3"].fillna(0.0).clip(lower=0.0)
    orders["volume_per_kg_used"] = lambda_v_per_kg
    return orders


def split_customer_into_batches(row: pd.Series) -> List[Dict[str, Any]]:
    """将单个客户需求拆为“满载批次 + 剩余批次”。

    拆分保留该客户货物的重量/体积比例。若重量约束更紧，则满载批次重量为 3000kg；
    若体积约束更紧，则满载批次体积为 13.5m^3。最后不足一个满载单位的剩余需求
    作为独立批次，后续可与其他客户剩余批次拼接到同一线路。
    """
    total_w = float(row["weight_kg"])
    total_v = float(row["volume_m3"])
    if total_w <= 1e-9 or total_v <= 1e-12:
        return []

    load_units = max(total_w / VEHICLE_WEIGHT_CAP, total_v / FUEL_VOLUME_CAP)
    if load_units <= 1.0 + 1e-9:
        return [
            {
                "weight_kg": total_w,
                "volume_m3": total_v,
                "batch_kind": "single",
                "limiting_dimension": "none",
                "capacity_fraction": load_units,
            }
        ]

    if total_w / VEHICLE_WEIGHT_CAP >= total_v / FUEL_VOLUME_CAP:
        limiting_dimension = "weight"
    else:
        limiting_dimension = "volume"

    full_count = int(math.floor(load_units + 1e-12))
    unit_w = total_w / load_units
    unit_v = total_v / load_units
    batches: List[Dict[str, Any]] = []
    remaining_w = total_w
    remaining_v = total_v

    for _ in range(full_count):
        bw = min(unit_w, remaining_w)
        bv = min(unit_v, remaining_v)
        if bw <= 1e-9 or bv <= 1e-12:
            break
        batches.append(
            {
                "weight_kg": bw,
                "volume_m3": bv,
                "batch_kind": "full",
                "limiting_dimension": limiting_dimension,
                "capacity_fraction": max(bw / VEHICLE_WEIGHT_CAP, bv / FUEL_VOLUME_CAP),
            }
        )
        remaining_w -= bw
        remaining_v -= bv

    if remaining_w > 1e-7 and remaining_v > 1e-10:
        batches.append(
            {
                "weight_kg": remaining_w,
                "volume_m3": remaining_v,
                "batch_kind": "remainder",
                "limiting_dimension": limiting_dimension,
                "capacity_fraction": max(remaining_w / VEHICLE_WEIGHT_CAP, remaining_v / FUEL_VOLUME_CAP),
            }
        )

    if batches:
        diff_w = total_w - sum(b["weight_kg"] for b in batches)
        diff_v = total_v - sum(b["volume_m3"] for b in batches)
        batches[-1]["weight_kg"] += diff_w
        batches[-1]["volume_m3"] += diff_v
        batches[-1]["capacity_fraction"] = max(
            batches[-1]["weight_kg"] / VEHICLE_WEIGHT_CAP,
            batches[-1]["volume_m3"] / FUEL_VOLUME_CAP,
        )
    return batches

def load_input_data(base_dir: Path) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """读取订单、距离矩阵、坐标与时间窗，并生成客户汇总表。"""
    order_path = base_dir / ORDER_FILE
    dist_path = base_dir / DIST_FILE
    coord_path = base_dir / COORD_FILE
    tw_path = base_dir / TW_FILE
    for p in [order_path, dist_path, coord_path, tw_path]:
        if not p.exists():
            raise FileNotFoundError(f"找不到输入文件: {p}")

    orders = pd.read_excel(order_path, sheet_name="完整订单数据")
    orders = orders.rename(columns={"目标客户编号": "customer_id", "重量": "weight_kg", "体积": "volume_m3"})
    orders["customer_id"] = orders["customer_id"].astype(int)
    orders = complete_order_weight_volume(orders)

    dist_raw = pd.read_excel(dist_path)
    first_col = dist_raw.columns[0]
    dist_raw = dist_raw.rename(columns={first_col: "from_id"})
    new_cols = ["from_id"] + [int(c) for c in dist_raw.columns[1:]]
    dist_raw.columns = new_cols
    dist_df = dist_raw.set_index("from_id").copy()
    dist_df.index = dist_df.index.astype(int)
    dist_df.columns = dist_df.columns.astype(int)

    coords = pd.read_excel(coord_path)
    coords = coords.rename(columns={"ID": "customer_id", "X (km)": "x_km", "Y (km)": "y_km", "类型": "type"})
    coords["customer_id"] = coords["customer_id"].astype(int)

    tw = pd.read_excel(tw_path)
    tw = tw.rename(columns={"客户编号": "customer_id", "开始时间": "tw_start", "结束时间": "tw_end"})
    tw["customer_id"] = tw["customer_id"].astype(int)
    tw["tw_start_min"] = tw["tw_start"].apply(hhmm_to_minute)
    tw["tw_end_min"] = tw["tw_end"].apply(hhmm_to_minute)

    agg = orders.groupby("customer_id", as_index=False).agg(
        weight_kg=("weight_kg", "sum"),
        volume_m3=("volume_m3", "sum"),
        order_count=("订单编号", "count") if "订单编号" in orders.columns else ("customer_id", "count"),
    )

    customer_df = coords.merge(agg, on="customer_id", how="left").merge(
        tw[["customer_id", "tw_start", "tw_end", "tw_start_min", "tw_end_min"]],
        on="customer_id",
        how="left",
    )
    customer_df["weight_kg"] = customer_df["weight_kg"].fillna(0.0)
    customer_df["volume_m3"] = customer_df["volume_m3"].fillna(0.0)
    customer_df["order_count"] = customer_df["order_count"].fillna(0).astype(int)

    return orders, dist_df, coords, tw, customer_df


def build_batches(customer_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """按“满载批次 + 剩余批次”的规则将客户需求拆分为批次。"""
    active = customer_df[
        (customer_df["customer_id"] != DEPOT_ID)
        & (customer_df["weight_kg"] > 1e-9)
        & (customer_df["volume_m3"] > 1e-12)
    ].copy()
    active = active.sort_values("customer_id").reset_index(drop=True)

    rows: List[Dict[str, Any]] = []
    for _, row in active.iterrows():
        pieces = split_customer_into_batches(row)
        split_count = len(pieces)
        for k, piece in enumerate(pieces, start=1):
            batch_weight = float(piece["weight_kg"])
            batch_volume = float(piece["volume_m3"])
            if batch_weight > VEHICLE_WEIGHT_CAP + 1e-7 or batch_volume > FUEL_VOLUME_CAP + 1e-7:
                raise ValueError(
                    f"客户 {int(row['customer_id'])} 第 {k} 个拆分批次超过容量："
                    f"{batch_weight:.4f}kg, {batch_volume:.4f}m3"
                )
            rows.append(
                {
                    "batch_id": len(rows) + 1,
                    "customer_id": int(row["customer_id"]),
                    "split_index": k,
                    "split_count": split_count,
                    "batch_kind": piece["batch_kind"],
                    "limiting_dimension": piece["limiting_dimension"],
                    "capacity_fraction": piece["capacity_fraction"],
                    "weight_kg": batch_weight,
                    "volume_m3": batch_volume,
                    "load_ratio_fuel": max(batch_weight / VEHICLE_WEIGHT_CAP, batch_volume / FUEL_VOLUME_CAP),
                    "x_km": float(row["x_km"]),
                    "y_km": float(row["y_km"]),
                    "tw_start_min": int(row["tw_start_min"]),
                    "tw_end_min": int(row["tw_end_min"]),
                    "tw_start": minute_to_hhmm(row["tw_start_min"]),
                    "tw_end": minute_to_hhmm(row["tw_end_min"]),
                }
            )
    batch_df = pd.DataFrame(rows)
    active = active.merge(
        batch_df.groupby("customer_id", as_index=False).agg(split_count=("batch_id", "count")),
        on="customer_id",
        how="left",
    )
    return active, batch_df


# =========================
# 4. 线路评价：时间窗、行驶距离、能耗
# =========================
def travel_time_minutes(distance_km: float, depart_minute: float) -> float:
    """给定出发时刻，按分段速度计算行驶时间。"""
    if distance_km <= 1e-12:
        return 0.0
    remain = float(distance_km)
    t = float(depart_minute)
    total_time = 0.0
    loop_guard = 0
    while remain > 1e-10:
        loop_guard += 1
        if loop_guard > 1000:
            raise RuntimeError("travel_time_minutes 出现异常循环。")
        matched = False
        for left, right, speed in TRAFFIC_INTERVALS:
            if left <= t < right:
                matched = True
                available_hours = (right - t) / 60.0
                possible_distance = speed * available_hours
                if possible_distance >= remain - 1e-10:
                    consume_min = remain / speed * 60.0
                    total_time += consume_min
                    t += consume_min
                    remain = 0.0
                else:
                    total_time += available_hours * 60.0
                    remain -= possible_distance
                    t = right
                break
        if not matched:
            # 若 t 小于第一段，推进到第一段；其他情况默认用最后一段速度。
            if t < TRAFFIC_INTERVALS[0][0]:
                t = TRAFFIC_INTERVALS[0][0]
            else:
                speed = TRAFFIC_INTERVALS[-1][2]
                consume_min = remain / speed * 60.0
                total_time += consume_min
                remain = 0.0
    return total_time


def route_load(seq: Iterable[int], batch_lookup: Dict[int, Dict[str, Any]]) -> Tuple[float, float]:
    w = sum(float(batch_lookup[bid]["weight_kg"]) for bid in seq)
    v = sum(float(batch_lookup[bid]["volume_m3"]) for bid in seq)
    return w, v


def route_is_capacity_feasible(seq: List[int], batch_lookup: Dict[int, Dict[str, Any]], volume_cap: float = FUEL_VOLUME_CAP) -> bool:
    if len(seq) > MAX_ROUTE_STOPS:
        return False
    w, v = route_load(seq, batch_lookup)
    return (w <= VEHICLE_WEIGHT_CAP + 1e-9) and (v <= volume_cap + 1e-9)


def estimate_route_departure(seq: List[int], batch_lookup: Dict[int, Dict[str, Any]], dist_df: pd.DataFrame) -> float:
    """根据首个停靠点时间窗估计线路出发时刻。"""
    first = batch_lookup[seq[0]]
    nominal_speed = 35.4
    dist_to_first = float(dist_df.loc[DEPOT_ID, int(first["customer_id"])])
    buffer_min = dist_to_first / nominal_speed * 60.0
    return max(float(START_OF_DAY), float(first["tw_start_min"]) - buffer_min)


def evaluate_route_fixed_start(
    seq: List[int],
    start_minute: float,
    batch_lookup: Dict[int, Dict[str, Any]],
    dist_df: pd.DataFrame,
) -> Dict[str, Any]:
    """给定线路顺序和固定出发时刻，计算该线路所有评价指标。"""
    if not seq:
        raise ValueError("空线路不能评价。")

    start_minute = max(float(START_OF_DAY), float(start_minute))
    t = start_minute
    load_w, load_v = route_load(seq, batch_lookup)
    remain_w, remain_v = load_w, load_v
    total_distance = 0.0
    total_wait = 0.0
    total_late = 0.0
    stops: List[Dict[str, Any]] = []
    legs: List[Dict[str, Any]] = []
    prev = DEPOT_ID

    for bid in seq:
        row = batch_lookup[bid]
        cid = int(row["customer_id"])
        dist = float(dist_df.loc[prev, cid])
        depart = t
        drive_min = travel_time_minutes(dist, depart)
        arrival = depart + drive_min
        wait = max(0.0, float(row["tw_start_min"]) - arrival)
        service_start = arrival + wait
        late = max(0.0, arrival - float(row["tw_end_min"]))
        leave = service_start + SERVICE_MIN

        total_distance += dist
        total_wait += wait
        total_late += late
        legs.append(
            {
                "from_id": prev,
                "to_id": cid,
                "distance_km": dist,
                "depart_min": depart,
                "drive_min": drive_min,
                "arrival_min": arrival,
                "remain_weight_kg": remain_w,
                "remain_volume_m3": remain_v,
            }
        )
        stops.append(
            {
                "batch_id": bid,
                "customer_id": cid,
                "arrival_min": arrival,
                "service_start_min": service_start,
                "leave_min": leave,
                "wait_min": wait,
                "late_min": late,
                "remain_weight_before_kg": remain_w,
                "remain_volume_before_m3": remain_v,
                "batch_weight_kg": float(row["weight_kg"]),
                "batch_volume_m3": float(row["volume_m3"]),
            }
        )
        remain_w -= float(row["weight_kg"])
        remain_v -= float(row["volume_m3"])
        prev = cid
        t = leave

    # 回仓
    back_dist = float(dist_df.loc[prev, DEPOT_ID])
    depart = t
    back_drive_min = travel_time_minutes(back_dist, depart)
    back_arrival = depart + back_drive_min
    total_distance += back_dist
    legs.append(
        {
            "from_id": prev,
            "to_id": DEPOT_ID,
            "distance_km": back_dist,
            "depart_min": depart,
            "drive_min": back_drive_min,
            "arrival_min": back_arrival,
            "remain_weight_kg": max(0.0, remain_w),
            "remain_volume_m3": max(0.0, remain_v),
        }
    )

    return {
        "start_min": start_minute,
        "end_min": back_arrival,
        "distance_km": total_distance,
        "wait_min": total_wait,
        "late_min": total_late,
        "load_weight_kg": load_w,
        "load_volume_m3": load_v,
        "stops": stops,
        "legs": legs,
    }


def evaluate_route(
    seq: List[int],
    batch_lookup: Dict[int, Dict[str, Any]],
    dist_df: pd.DataFrame,
    start_minute: Optional[float] = None,
    auto_shift: bool = True,
    final_search: bool = False,
) -> Dict[str, Any]:
    """自动选择或微调线路出发时刻，并评价线路。"""
    if start_minute is None:
        start_minute = estimate_route_departure(seq, batch_lookup, dist_df)
    eva = evaluate_route_fixed_start(seq, start_minute, batch_lookup, dist_df)

    # 向后平移，尽量减少等待时间，同时不增加迟到。
    if auto_shift:
        for _ in range(20):
            waits = [s["wait_min"] for s in eva["stops"] if s["wait_min"] > 1e-8]
            if not waits:
                break
            shift = min(waits)
            cand = evaluate_route_fixed_start(seq, eva["start_min"] + shift, batch_lookup, dist_df)
            if cand["late_min"] <= eva["late_min"] + 1e-8 and cand["wait_min"] < eva["wait_min"] - 1e-8:
                eva = cand
            else:
                break

    # 最终输出前做小范围离散搜索，进一步减小时间窗成本。
    if final_search:
        center = eva["start_min"]
        lo = max(START_OF_DAY, center - 90)
        hi = min(END_OF_DAY, center + 90)
        candidates = list(np.arange(lo, hi + 1e-9, 5.0)) + [center]
        best = eva
        best_obj = time_window_cost(best) + 0.02 * best["distance_km"] + max(0.0, best["end_min"] - END_OF_DAY) * CROSS_DAY_PENALTY_PER_MIN
        for st in candidates:
            cand = evaluate_route_fixed_start(seq, st, batch_lookup, dist_df)
            obj = time_window_cost(cand) + 0.02 * cand["distance_km"] + max(0.0, cand["end_min"] - END_OF_DAY) * CROSS_DAY_PENALTY_PER_MIN
            if obj < best_obj - 1e-9:
                best = cand
                best_obj = obj
        eva = best
    return eva


def time_window_cost(route_eval: Dict[str, Any]) -> float:
    return route_eval["wait_min"] / 60.0 * WAIT_COST_PER_HOUR + route_eval["late_min"] / 60.0 * LATE_COST_PER_HOUR


def compute_route_energy_cost(route_eval: Dict[str, Any], vehicle_type: str) -> Tuple[float, float, float]:
    """
    返回：能源费用、碳排费用、能源消耗量（燃油为 L，新能源为 kWh）。
    vehicle_type: 'fuel' 或 'ev'
    """
    energy_cost = 0.0
    carbon_cost = 0.0
    total_consumption = 0.0
    for leg in route_eval["legs"]:
        dist = float(leg["distance_km"])
        drive_min = float(leg["drive_min"])
        if dist <= 1e-12:
            continue
        avg_speed = dist / (drive_min / 60.0) if drive_min > 1e-12 else 35.4
        remain_w = max(0.0, float(leg["remain_weight_kg"]))
        remain_v = max(0.0, float(leg["remain_volume_m3"]))
        if vehicle_type == "fuel":
            load_ratio = max(remain_w / VEHICLE_WEIGHT_CAP, remain_v / FUEL_VOLUME_CAP)
            load_ratio = min(1.0, max(0.0, load_ratio))
            consumption_per_100 = fuel_per_100km(avg_speed) * (1.0 + 0.40 * load_ratio)
            consumption = dist / 100.0 * consumption_per_100
            total_consumption += consumption
            energy_cost += consumption * FUEL_PRICE
            carbon_cost += consumption * FUEL_CARBON_FACTOR * CARBON_PRICE
        elif vehicle_type == "ev":
            load_ratio = max(remain_w / VEHICLE_WEIGHT_CAP, remain_v / EV_VOLUME_CAP)
            load_ratio = min(1.0, max(0.0, load_ratio))
            consumption_per_100 = electric_per_100km(avg_speed) * (1.0 + 0.35 * load_ratio)
            consumption = dist / 100.0 * consumption_per_100
            total_consumption += consumption
            energy_cost += consumption * ELECTRIC_PRICE
            carbon_cost += consumption * ELECTRIC_CARBON_FACTOR * CARBON_PRICE
        else:
            raise ValueError("vehicle_type 必须为 'fuel' 或 'ev'")
    return energy_cost, carbon_cost, total_consumption


def route_construct_score(route_eval: Dict[str, Any]) -> float:
    """构造/局部搜索阶段使用的单线路近似成本。"""
    e, c, _ = compute_route_energy_cost(route_eval, "fuel")
    return (
        APPROX_FIXED_COST_PER_ROUTE
        + e
        + c
        + time_window_cost(route_eval)
        + max(0.0, route_eval["end_min"] - END_OF_DAY) * CROSS_DAY_PENALTY_PER_MIN
    )


# =========================
# 5. 启发式构造与局部搜索
# =========================
def batch_sort_order(batch_df: pd.DataFrame) -> List[int]:
    """批次排序：先按时间窗，再优先安排高载荷批次。"""
    tmp = batch_df.copy()
    tmp["tw_bucket"] = (tmp["tw_start_min"] // 60).astype(int)
    tmp["angle"] = np.arctan2(tmp["y_km"] - 20.0, tmp["x_km"] - 20.0)
    tmp = tmp.sort_values(
        ["tw_start_min", "tw_end_min", "load_ratio_fuel", "angle", "customer_id"],
        ascending=[True, True, False, True, True],
    )
    return tmp["batch_id"].astype(int).tolist()


def build_initial_routes(batch_df: pd.DataFrame, batch_lookup: Dict[int, Dict[str, Any]], dist_df: pd.DataFrame) -> List[List[int]]:
    """顺序插入构造初始线路。"""
    routes: List[List[int]] = []
    route_evals: List[Dict[str, Any]] = []
    route_scores: List[float] = []

    for bid in batch_sort_order(batch_df):
        singleton = [bid]
        singleton_eval = evaluate_route(singleton, batch_lookup, dist_df)
        singleton_score = route_construct_score(singleton_eval)

        best: Optional[Tuple[float, int, List[int], Dict[str, Any], float]] = None
        for ridx, seq in enumerate(routes):
            if len(seq) >= MAX_ROUTE_STOPS:
                continue
            current_w, current_v = route_load(seq, batch_lookup)
            add_w = float(batch_lookup[bid]["weight_kg"])
            add_v = float(batch_lookup[bid]["volume_m3"])
            if current_w + add_w > VEHICLE_WEIGHT_CAP + 1e-9:
                continue
            if current_v + add_v > FUEL_VOLUME_CAP + 1e-9:
                continue

            old_score = route_scores[ridx]
            for pos in range(len(seq) + 1):
                cand_seq = seq[:pos] + [bid] + seq[pos:]
                cand_eval = evaluate_route(cand_seq, batch_lookup, dist_df)
                cand_score = route_construct_score(cand_eval)
                delta = cand_score - old_score
                if best is None or delta < best[0]:
                    best = (delta, ridx, cand_seq, cand_eval, cand_score)

        # 只有插入增量小于新开一条线路，才插入已有线路。
        if best is not None and best[0] < singleton_score:
            _, ridx, cand_seq, cand_eval, cand_score = best
            routes[ridx] = cand_seq
            route_evals[ridx] = cand_eval
            route_scores[ridx] = cand_score
        else:
            routes.append(singleton)
            route_evals.append(singleton_eval)
            route_scores.append(singleton_score)

    return routes


def refine_single_route(seq: List[int], batch_lookup: Dict[int, Dict[str, Any]], dist_df: pd.DataFrame) -> Tuple[List[int], Dict[str, Any]]:
    """路线内部排序优化。MAX_ROUTE_STOPS=5 时可枚举全排列。"""
    best_seq = list(seq)
    best_eval = evaluate_route(best_seq, batch_lookup, dist_df)
    best_score = route_construct_score(best_eval)

    if len(seq) <= MAX_ROUTE_STOPS:
        for perm in itertools.permutations(seq):
            cand_seq = list(perm)
            cand_eval = evaluate_route(cand_seq, batch_lookup, dist_df)
            cand_score = route_construct_score(cand_eval)
            if cand_score < best_score - 1e-9:
                best_seq = cand_seq
                best_eval = cand_eval
                best_score = cand_score
    return best_seq, best_eval


def refine_routes_order(routes: List[List[int]], batch_lookup: Dict[int, Dict[str, Any]], dist_df: pd.DataFrame) -> Tuple[List[List[int]], List[Dict[str, Any]], List[float]]:
    new_routes = []
    route_evals = []
    route_scores = []
    for seq in routes:
        if not seq:
            continue
        s, e = refine_single_route(seq, batch_lookup, dist_df)
        new_routes.append(s)
        route_evals.append(e)
        route_scores.append(route_construct_score(e))
    return new_routes, route_evals, route_scores


def local_relocation_search(
    routes: List[List[int]],
    batch_lookup: Dict[int, Dict[str, Any]],
    dist_df: pd.DataFrame,
    max_iter: int = 4,
) -> List[List[int]]:
    """跨线路单批次重定位搜索，主要用于降低路线数、距离和时间窗罚金。"""
    routes, evals, scores = refine_routes_order(routes, batch_lookup, dist_df)

    for _ in range(max_iter):
        best_move: Optional[Dict[str, Any]] = None
        current_total = sum(scores)
        n = len(routes)
        # 尝试把 i 线路中的一个批次移到 j 线路，或移成新单点线路。
        for i in range(n):
            if not routes[i]:
                continue
            for pos_i, bid in enumerate(routes[i]):
                seq_i_removed = routes[i][:pos_i] + routes[i][pos_i + 1 :]
                if seq_i_removed:
                    eval_i_removed = evaluate_route(seq_i_removed, batch_lookup, dist_df)
                    score_i_removed = route_construct_score(eval_i_removed)
                else:
                    eval_i_removed = None
                    score_i_removed = 0.0

                # 与新开线路比较：一般不会改进，但保留该动作以允许拆散差路线。
                singleton_eval = evaluate_route([bid], batch_lookup, dist_df)
                singleton_score = route_construct_score(singleton_eval)
                delta_new = score_i_removed + singleton_score - scores[i]
                if delta_new < -1e-6 and (best_move is None or delta_new < best_move["delta"]):
                    best_move = {
                        "delta": delta_new,
                        "from": i,
                        "to": None,
                        "bid": bid,
                        "seq_i": seq_i_removed,
                        "eval_i": eval_i_removed,
                        "score_i": score_i_removed,
                        "seq_j": [bid],
                        "eval_j": singleton_eval,
                        "score_j": singleton_score,
                    }

                for j in range(n):
                    if i == j:
                        continue
                    if len(routes[j]) >= MAX_ROUTE_STOPS:
                        continue
                    wj, vj = route_load(routes[j], batch_lookup)
                    add_w = float(batch_lookup[bid]["weight_kg"])
                    add_v = float(batch_lookup[bid]["volume_m3"])
                    if wj + add_w > VEHICLE_WEIGHT_CAP + 1e-9 or vj + add_v > FUEL_VOLUME_CAP + 1e-9:
                        continue
                    for pos_j in range(len(routes[j]) + 1):
                        seq_j = routes[j][:pos_j] + [bid] + routes[j][pos_j:]
                        cand_eval_j = evaluate_route(seq_j, batch_lookup, dist_df)
                        cand_score_j = route_construct_score(cand_eval_j)
                        delta = score_i_removed + cand_score_j - scores[i] - scores[j]
                        if delta < -1e-6 and (best_move is None or delta < best_move["delta"]):
                            best_move = {
                                "delta": delta,
                                "from": i,
                                "to": j,
                                "bid": bid,
                                "seq_i": seq_i_removed,
                                "eval_i": eval_i_removed,
                                "score_i": score_i_removed,
                                "seq_j": seq_j,
                                "eval_j": cand_eval_j,
                                "score_j": cand_score_j,
                            }

        if best_move is None:
            break

        i = best_move["from"]
        j = best_move["to"]
        if j is None:
            # 移为新线路
            routes[i] = best_move["seq_i"]
            if best_move["seq_i"]:
                evals[i] = best_move["eval_i"]
                scores[i] = best_move["score_i"]
            else:
                routes.pop(i)
                evals.pop(i)
                scores.pop(i)
            routes.append(best_move["seq_j"])
            evals.append(best_move["eval_j"])
            scores.append(best_move["score_j"])
        else:
            routes[i] = best_move["seq_i"]
            routes[j] = best_move["seq_j"]
            if best_move["seq_i"]:
                evals[i] = best_move["eval_i"]
                scores[i] = best_move["score_i"]
            else:
                # 删除空线路时需要修正 j 的索引
                routes.pop(i)
                evals.pop(i)
                scores.pop(i)
                if j > i:
                    j -= 1
            evals[j] = best_move["eval_j"]
            scores[j] = best_move["score_j"]

        # 每次重定位后重新优化线路内部顺序，防止局部顺序恶化。
        routes, evals, scores = refine_routes_order(routes, batch_lookup, dist_df)
        new_total = sum(scores)
        if new_total > current_total + 1e-4:
            # 理论上不会出现，若出现则停止。
            break
    return routes


# =========================
# 6. 新能源车选择与实体车辆复用
# =========================
def choose_ev_route_ids_by_min_cost_flow(positive_df: pd.DataFrame, ev_limit: int) -> set[int]:
    """纯 Python 最小费用流：在任意时刻最多 ev_limit 条 EV 线路的约束下最大化节省额。"""
    if positive_df.empty or ev_limit <= 0:
        return set()

    times = sorted(set(positive_df["start_min"].tolist() + positive_df["end_min"].tolist()))
    time_index = {t: i for i, t in enumerate(times)}
    n = len(times)
    graph: List[List[Dict[str, Any]]] = [[] for _ in range(n)]
    route_edges: List[Tuple[int, int, int]] = []

    def add_edge(u: int, v: int, cap: int, cost: float, route_id: Optional[int] = None) -> None:
        fwd = {"to": v, "rev": len(graph[v]), "cap": cap, "cost": cost, "route_id": route_id}
        rev = {"to": u, "rev": len(graph[u]), "cap": 0, "cost": -cost, "route_id": None}
        graph[u].append(fwd)
        graph[v].append(rev)
        if route_id is not None:
            route_edges.append((u, len(graph[u]) - 1, int(route_id)))

    for i in range(n - 1):
        add_edge(i, i + 1, ev_limit, 0.0)

    scale = 10000.0
    for row in positive_df.to_dict("records"):
        u = time_index[row["start_min"]]
        v = time_index[row["end_min"]]
        if u != v:
            add_edge(u, v, 1, -float(row["ev_saving"]) * scale, int(row["route_id"]))

    source, sink = 0, n - 1
    selected_flow = 0
    while selected_flow < ev_limit:
        inf = float("inf")
        dist = [inf] * n
        prev_node = [-1] * n
        prev_edge = [-1] * n
        dist[source] = 0.0

        # Bellman-Ford 可直接处理负费用边；节点数量很小，计算量可接受。
        for _ in range(n - 1):
            updated = False
            for u in range(n):
                if dist[u] == inf:
                    continue
                for ei, edge in enumerate(graph[u]):
                    if edge["cap"] <= 0:
                        continue
                    v = edge["to"]
                    nd = dist[u] + edge["cost"]
                    if nd < dist[v] - 1e-9:
                        dist[v] = nd
                        prev_node[v] = u
                        prev_edge[v] = ei
                        updated = True
            if not updated:
                break

        # 没有负费用增广路时，继续增广只会走时间边，不能带来节约。
        if dist[sink] == inf or dist[sink] >= -1e-9:
            break

        add_flow = ev_limit - selected_flow
        v = sink
        while v != source:
            u = prev_node[v]
            ei = prev_edge[v]
            if u < 0 or ei < 0:
                add_flow = 0
                break
            add_flow = min(add_flow, int(graph[u][ei]["cap"]))
            v = u
        if add_flow <= 0:
            break

        v = sink
        while v != source:
            u = prev_node[v]
            ei = prev_edge[v]
            rev_i = graph[u][ei]["rev"]
            graph[u][ei]["cap"] -= add_flow
            graph[v][rev_i]["cap"] += add_flow
            v = u
        selected_flow += add_flow

    selected_ids = set()
    for u, ei, route_id in route_edges:
        # 正向边原始容量为 1；若剩余容量为 0，说明最终有 1 单位流选择了该线路。
        if graph[u][ei]["cap"] == 0:
            selected_ids.add(route_id)
    return selected_ids


def select_ev_routes(route_evals: List[Dict[str, Any]], ev_limit: int = EV_FLEET_LIMIT) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """选择新能源车执行线路，并保证任意时刻新能源车占用数不超过上限。"""
    rows = []
    for rid, eva in enumerate(route_evals, start=1):
        fuel_energy, fuel_carbon, fuel_cons = compute_route_energy_cost(eva, "fuel")
        ev_energy, ev_carbon, ev_cons = compute_route_energy_cost(eva, "ev")
        fuel_total = fuel_energy + fuel_carbon
        ev_total = ev_energy + ev_carbon
        rows.append(
            {
                "route_id": rid,
                "start_min": eva["start_min"],
                "end_min": eva["end_min"],
                "fuel_energy_cost": fuel_energy,
                "fuel_carbon_cost": fuel_carbon,
                "fuel_consumption_l": fuel_cons,
                "ev_energy_cost": ev_energy,
                "ev_carbon_cost": ev_carbon,
                "ev_consumption_kwh": ev_cons,
                "fuel_variable_cost": fuel_total,
                "ev_variable_cost": ev_total,
                "ev_saving": fuel_total - ev_total,
            }
        )
    gain_df = pd.DataFrame(rows)

    positive = gain_df[gain_df["ev_saving"] > 1e-9].copy()
    selected_ev_ids = choose_ev_route_ids_by_min_cost_flow(positive, ev_limit)

    assign_rows = []
    for row in gain_df.to_dict("records"):
        if int(row["route_id"]) in selected_ev_ids:
            assign_rows.append(
                {
                    "route_id": int(row["route_id"]),
                    "vehicle_type": "新能源3000kg",
                    "energy_cost": row["ev_energy_cost"],
                    "carbon_cost": row["ev_carbon_cost"],
                    "energy_consumption": row["ev_consumption_kwh"],
                    "saving_vs_fuel": row["ev_saving"],
                }
            )
        else:
            assign_rows.append(
                {
                    "route_id": int(row["route_id"]),
                    "vehicle_type": "燃油3000kg",
                    "energy_cost": row["fuel_energy_cost"],
                    "carbon_cost": row["fuel_carbon_cost"],
                    "energy_consumption": row["fuel_consumption_l"],
                    "saving_vs_fuel": 0.0,
                }
            )
    assign_df = pd.DataFrame(assign_rows)
    return gain_df, assign_df


def max_overlap(interval_rows: List[Dict[str, Any]]) -> int:
    events = []
    for row in interval_rows:
        events.append((float(row["start_min"]), 1))
        events.append((float(row["end_min"]), -1))
    # 同一时刻先释放车辆再占用车辆。
    events.sort(key=lambda x: (x[0], x[1]))
    cur = 0
    ans = 0
    for _, d in events:
        cur += d
        ans = max(ans, cur)
    return ans


def assign_physical_vehicles(route_summary_df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """在同一车辆类型内按时间区间复用实体车辆。"""
    rows = []
    vehicle_count: Dict[str, int] = {}
    for vehicle_type, sub in route_summary_df.groupby("vehicle_type"):
        heap: List[Tuple[float, int]] = []
        next_no = 1
        first_route: Dict[int, int] = {}
        for row in sub.sort_values(["start_min", "end_min", "route_id"]).to_dict("records"):
            if heap and heap[0][0] <= float(row["start_min"]) + 1e-9:
                _, vehicle_no = heapq.heappop(heap)
            else:
                vehicle_no = next_no
                first_route[vehicle_no] = int(row["route_id"])
                next_no += 1
            heapq.heappush(heap, (float(row["end_min"]), vehicle_no))
            rows.append(
                {
                    "route_id": int(row["route_id"]),
                    "vehicle_type": vehicle_type,
                    "vehicle_id": f"{vehicle_type}-{vehicle_no:02d}",
                    "is_new_vehicle": int(first_route[vehicle_no] == int(row["route_id"])),
                }
            )
        vehicle_count[vehicle_type] = next_no - 1
    return pd.DataFrame(rows), vehicle_count


# =========================
# 7. 结果表、图片与 Excel 输出
# =========================
def build_result_tables(
    routes: List[List[int]],
    route_evals: List[Dict[str, Any]],
    energy_assign_df: pd.DataFrame,
    batch_lookup: Dict[int, Dict[str, Any]],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    route_rows = []
    stop_rows = []
    leg_rows = []
    assign_lookup = energy_assign_df.set_index("route_id").to_dict("index")

    for rid, (seq, eva) in enumerate(zip(routes, route_evals), start=1):
        assign = assign_lookup[rid]
        vehicle_type = assign["vehicle_type"]
        energy_cost = float(assign["energy_cost"])
        carbon_cost = float(assign["carbon_cost"])
        wait_cost = eva["wait_min"] / 60.0 * WAIT_COST_PER_HOUR
        late_cost = eva["late_min"] / 60.0 * LATE_COST_PER_HOUR
        customer_seq = [int(batch_lookup[x]["customer_id"]) for x in seq]
        route_rows.append(
            {
                "route_id": rid,
                "vehicle_type": vehicle_type,
                "batch_count": len(seq),
                "customer_sequence": "0-" + "-".join(map(str, customer_seq)) + "-0",
                "batch_sequence": "|".join(map(str, seq)),
                "load_weight_kg": eva["load_weight_kg"],
                "load_volume_m3": eva["load_volume_m3"],
                "start_min": eva["start_min"],
                "end_min": eva["end_min"],
                "start_time": minute_to_hhmm(eva["start_min"]),
                "end_time": minute_to_hhmm(eva["end_min"]),
                "distance_km": eva["distance_km"],
                "wait_min": eva["wait_min"],
                "late_min": eva["late_min"],
                "energy_cost": energy_cost,
                "carbon_cost": carbon_cost,
                "energy_consumption": float(assign["energy_consumption"]),
                "saving_vs_fuel": float(assign["saving_vs_fuel"]),
                "wait_cost": wait_cost,
                "late_cost": late_cost,
            }
        )
        for order_no, stop in enumerate(eva["stops"], start=1):
            stop_rows.append(
                {
                    "route_id": rid,
                    "stop_order": order_no,
                    "batch_id": stop["batch_id"],
                    "customer_id": stop["customer_id"],
                    "vehicle_type": vehicle_type,
                    "arrival_min": stop["arrival_min"],
                    "service_start_min": stop["service_start_min"],
                    "leave_min": stop["leave_min"],
                    "arrival_time": minute_to_hhmm(stop["arrival_min"]),
                    "service_start_time": minute_to_hhmm(stop["service_start_min"]),
                    "leave_time": minute_to_hhmm(stop["leave_min"]),
                    "wait_min": stop["wait_min"],
                    "late_min": stop["late_min"],
                    "batch_weight_kg": stop["batch_weight_kg"],
                    "batch_volume_m3": stop["batch_volume_m3"],
                    "remain_weight_before_kg": stop["remain_weight_before_kg"],
                    "remain_volume_before_m3": stop["remain_volume_before_m3"],
                }
            )
        for leg_no, leg in enumerate(eva["legs"], start=1):
            leg_rows.append(
                {
                    "route_id": rid,
                    "leg_order": leg_no,
                    "vehicle_type": vehicle_type,
                    "from_id": leg["from_id"],
                    "to_id": leg["to_id"],
                    "distance_km": leg["distance_km"],
                    "depart_min": leg["depart_min"],
                    "arrival_min": leg["arrival_min"],
                    "depart_time": minute_to_hhmm(leg["depart_min"]),
                    "arrival_time": minute_to_hhmm(leg["arrival_min"]),
                    "drive_min": leg["drive_min"],
                    "avg_speed_kmh": leg["distance_km"] / (leg["drive_min"] / 60.0) if leg["drive_min"] > 1e-12 else 0.0,
                    "remain_weight_kg": leg["remain_weight_kg"],
                    "remain_volume_m3": leg["remain_volume_m3"],
                }
            )
    return pd.DataFrame(route_rows), pd.DataFrame(stop_rows), pd.DataFrame(leg_rows)


def add_vehicle_costs(route_summary_df: pd.DataFrame, vehicle_chain_df: pd.DataFrame) -> pd.DataFrame:
    merged = route_summary_df.merge(vehicle_chain_df, on=["route_id", "vehicle_type"], how="left")
    merged["fixed_cost"] = np.where(merged["is_new_vehicle"] == 1, FIXED_VEHICLE_COST, 0.0)
    merged["total_cost"] = (
        merged["fixed_cost"]
        + merged["energy_cost"]
        + merged["carbon_cost"]
        + merged["wait_cost"]
        + merged["late_cost"]
    )
    return merged.sort_values("route_id").reset_index(drop=True)


def build_vehicle_usage(route_summary_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for vid, sub in route_summary_df.groupby("vehicle_id"):
        rows.append(
            {
                "vehicle_id": vid,
                "vehicle_type": sub["vehicle_type"].iloc[0],
                "route_count": len(sub),
                "first_start_time": minute_to_hhmm(sub["start_min"].min()),
                "last_end_time": minute_to_hhmm(sub["end_min"].max()),
                "first_start_min": sub["start_min"].min(),
                "last_end_min": sub["end_min"].max(),
                "distance_km": sub["distance_km"].sum(),
                "total_cost": sub["total_cost"].sum(),
            }
        )
    return pd.DataFrame(rows).sort_values(["vehicle_type", "vehicle_id"]).reset_index(drop=True)


def plot_transport_flow(route_summary_df: pd.DataFrame, stop_schedule_df: pd.DataFrame, coords_df: pd.DataFrame, out_path: Path) -> None:
    set_matplotlib_font()
    coord = coords_df[["customer_id", "x_km", "y_km"]].drop_duplicates().set_index("customer_id").to_dict("index")
    fig, ax = plt.subplots(figsize=(10.5, 10.0))

    # 客户点
    customers = coords_df[coords_df["customer_id"] != DEPOT_ID]
    ax.scatter(customers["x_km"], customers["y_km"], s=18, alpha=0.58, label="Customer")
    ax.scatter(coord[DEPOT_ID]["x_km"], coord[DEPOT_ID]["y_km"], marker="*", s=260, label="Depot")

    type_lookup = route_summary_df.set_index("route_id")["vehicle_type"].to_dict()
    color_lookup = {"燃油3000kg": "tab:blue", "新能源3000kg": "tab:green"}
    label_done = set()
    for rid, sub in stop_schedule_df.groupby("route_id"):
        sub = sub.sort_values("stop_order")
        seq = [DEPOT_ID] + sub["customer_id"].astype(int).tolist() + [DEPOT_ID]
        xs = [coord[i]["x_km"] for i in seq]
        ys = [coord[i]["y_km"] for i in seq]
        vt = type_lookup[int(rid)]
        label = "Fuel route" if vt == "燃油3000kg" else "EV route"
        show_label = label if label not in label_done else None
        label_done.add(label)
        ax.plot(xs, ys, linewidth=0.9, alpha=0.30, color=color_lookup[vt], label=show_label)
        # 在第一段画轻量箭头表示方向
        if len(xs) >= 2:
            ax.annotate(
                "",
                xy=(xs[1], ys[1]),
                xytext=(xs[0], ys[0]),
                arrowprops=dict(arrowstyle="->", lw=0.6, alpha=0.22, color=color_lookup[vt]),
            )

    ax.set_title("Q1 Transport Flow Map")
    ax.set_xlabel("X (km)")
    ax.set_ylabel("Y (km)")
    ax.grid(alpha=0.25, linestyle="--")
    ax.legend(loc="best")
    ax.set_aspect("equal", adjustable="box")
    plt.tight_layout()
    fig.savefig(out_path, bbox_inches="tight")
    plt.close(fig)


def plot_vehicle_gantt(route_summary_df: pd.DataFrame, out_path: Path, title: str = "Q1 Vehicle Reuse Gantt") -> None:
    set_matplotlib_font()
    df = route_summary_df.sort_values(["vehicle_type", "vehicle_id", "start_min"]).copy()

    def display_vehicle_id(raw_id: str, vehicle_type: str) -> str:
        suffix = str(raw_id).split("-")[-1]
        return ("EV-" if vehicle_type == "新能源3000kg" else "F-") + suffix

    df["vehicle_label"] = [display_vehicle_id(v, t) for v, t in zip(df["vehicle_id"], df["vehicle_type"])]
    vehicle_order = df["vehicle_label"].drop_duplicates().tolist()
    y_map = {v: i for i, v in enumerate(vehicle_order)}
    color_lookup = {"燃油3000kg": "tab:blue", "新能源3000kg": "tab:green"}
    fig_h = max(7, 0.22 * len(vehicle_order))
    fig, ax = plt.subplots(figsize=(13, fig_h))
    for _, row in df.iterrows():
        y = y_map[row["vehicle_label"]]
        ax.barh(
            y=y,
            width=row["end_min"] - row["start_min"],
            left=row["start_min"],
            height=0.55,
            color=color_lookup[row["vehicle_type"]],
            alpha=0.78,
        )
        ax.text(row["start_min"], y, f"R{int(row['route_id'])}", va="center", ha="left", fontsize=6)
    xticks = list(range(8 * 60, 23 * 60 + 1, 60))
    ax.set_xticks(xticks)
    ax.set_xticklabels([minute_to_hhmm(x) for x in xticks], rotation=45)
    ax.set_yticks(list(y_map.values()))
    ax.set_yticklabels(vehicle_order, fontsize=6)
    ax.set_title(str(title))
    ax.set_xlabel("Time")
    ax.set_ylabel("Vehicle")
    ax.grid(axis="x", alpha=0.25, linestyle="--")
    plt.tight_layout()
    fig.savefig(out_path, bbox_inches="tight")
    plt.close(fig)


def plot_cost_breakdown(route_summary_df: pd.DataFrame, out_path: Path) -> None:
    set_matplotlib_font()
    data = pd.Series(
        {
            "Fixed": route_summary_df["fixed_cost"].sum(),
            "Energy": route_summary_df["energy_cost"].sum(),
            "Carbon": route_summary_df["carbon_cost"].sum(),
            "Waiting": route_summary_df["wait_cost"].sum(),
            "Late": route_summary_df["late_cost"].sum(),
        }
    )
    fig, ax = plt.subplots(figsize=(8.8, 5.2))
    bars = ax.bar(data.index, data.values)
    ax.set_title("Q1 Cost Breakdown")
    ax.set_ylabel("Cost (CNY)")
    ax.grid(axis="y", alpha=0.25, linestyle="--")
    for bar in bars:
        y = bar.get_height()
        ax.text(bar.get_x() + bar.get_width() / 2, y, f"{y:.1f}", ha="center", va="bottom", fontsize=8)
    plt.tight_layout()
    fig.savefig(out_path, bbox_inches="tight")
    plt.close(fig)


def export_excel(
    result_path: Path,
    summary_df: pd.DataFrame,
    customer_summary_df: pd.DataFrame,
    batch_df: pd.DataFrame,
    route_summary_df: pd.DataFrame,
    stop_schedule_df: pd.DataFrame,
    leg_df: pd.DataFrame,
    vehicle_chain_df: pd.DataFrame,
    vehicle_usage_df: pd.DataFrame,
    energy_gain_df: pd.DataFrame,
    image_paths: List[Path],
) -> None:
    # 为易读性，Excel 采用中文表头；脚本内部仍使用英文变量名便于复现。
    route_cn = {
        "route_id": "线路编号", "vehicle_type": "车辆类型", "batch_count": "批次数",
        "customer_sequence": "客户序列", "batch_sequence": "批次序列",
        "load_weight_kg": "装载重量kg", "load_volume_m3": "装载体积m3",
        "start_min": "出发分钟", "end_min": "返回分钟", "start_time": "出发时间", "end_time": "返回时间",
        "distance_km": "行驶距离km", "wait_min": "等待分钟", "late_min": "迟到分钟",
        "energy_cost": "能源成本元", "carbon_cost": "碳排成本元", "energy_consumption": "能源消耗量",
        "saving_vs_fuel": "较燃油节省元", "wait_cost": "等待成本元", "late_cost": "迟到成本元",
        "vehicle_id": "实体车辆编号", "is_new_vehicle": "是否新增车辆", "fixed_cost": "固定成本元", "total_cost": "线路总成本元",
    }
    stop_cn = {
        "route_id": "线路编号", "stop_order": "停靠序号", "batch_id": "批次编号", "customer_id": "客户编号",
        "vehicle_type": "车辆类型", "arrival_min": "到达分钟", "service_start_min": "开始服务分钟", "leave_min": "离开分钟",
        "arrival_time": "到达时间", "service_start_time": "开始服务时间", "leave_time": "离开时间",
        "wait_min": "等待分钟", "late_min": "迟到分钟", "batch_weight_kg": "批次重量kg", "batch_volume_m3": "批次体积m3",
        "remain_weight_before_kg": "到站前剩余重量kg", "remain_volume_before_m3": "到站前剩余体积m3",
    }
    leg_cn = {
        "route_id": "线路编号", "leg_order": "路段序号", "vehicle_type": "车辆类型", "from_id": "起点", "to_id": "终点",
        "distance_km": "距离km", "depart_min": "出发分钟", "arrival_min": "到达分钟", "depart_time": "出发时间",
        "arrival_time": "到达时间", "drive_min": "行驶分钟", "avg_speed_kmh": "平均速度kmh",
        "remain_weight_kg": "路段载重kg", "remain_volume_m3": "路段载容m3",
    }
    vehicle_usage_cn = {
        "vehicle_id": "实体车辆编号", "vehicle_type": "车辆类型", "route_count": "承担线路数",
        "first_start_time": "首次出发时间", "last_end_time": "最后返回时间", "first_start_min": "首次出发分钟",
        "last_end_min": "最后返回分钟", "distance_km": "累计行驶距离km", "total_cost": "累计成本元",
    }
    vehicle_chain_cn = {
        "route_id": "线路编号", "vehicle_type": "车辆类型", "vehicle_id": "实体车辆编号", "is_new_vehicle": "是否新增车辆",
    }
    energy_cn = {
        "route_id": "线路编号", "start_min": "出发分钟", "end_min": "返回分钟",
        "fuel_energy_cost": "燃油能源成本元", "fuel_carbon_cost": "燃油碳排成本元", "fuel_consumption_l": "燃油消耗L",
        "ev_energy_cost": "新能源能源成本元", "ev_carbon_cost": "新能源碳排成本元", "ev_consumption_kwh": "新能源电耗kWh",
        "fuel_variable_cost": "燃油变量成本元", "ev_variable_cost": "新能源变量成本元", "ev_saving": "新能源节省元",
    }
    batch_cn = {
        "batch_id": "批次编号", "customer_id": "客户编号", "split_index": "拆分序号", "split_count": "拆分总数",
        "batch_kind": "批次类型", "limiting_dimension": "限制维度", "capacity_fraction": "容量占用率",
        "weight_kg": "批次重量kg", "volume_m3": "批次体积m3", "load_ratio_fuel": "燃油容量占用率",
        "x_km": "X坐标km", "y_km": "Y坐标km", "tw_start_min": "时间窗开始分钟", "tw_end_min": "时间窗结束分钟",
        "tw_start": "时间窗开始", "tw_end": "时间窗结束",
    }
    customer_cn = {
        "customer_id": "客户编号", "order_count": "订单数", "weight_kg": "总重量kg", "volume_m3": "总体积m3",
        "split_count": "拆分批次数", "x_km": "X坐标km", "y_km": "Y坐标km",
        "tw_start_min": "时间窗开始分钟", "tw_end_min": "时间窗结束分钟", "tw_start": "时间窗开始", "tw_end": "时间窗结束",
    }

    with pd.ExcelWriter(result_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="汇总结果", index=False)
        route_summary_df.rename(columns=route_cn).to_excel(writer, sheet_name="线路汇总", index=False)
        stop_schedule_df.rename(columns=stop_cn).to_excel(writer, sheet_name="停靠时刻表", index=False)
        leg_df.rename(columns=leg_cn).to_excel(writer, sheet_name="路段明细", index=False)
        vehicle_usage_df.rename(columns=vehicle_usage_cn).to_excel(writer, sheet_name="实体车辆使用", index=False)
        vehicle_chain_df.rename(columns=vehicle_chain_cn).to_excel(writer, sheet_name="线路-车辆映射", index=False)
        energy_gain_df.rename(columns=energy_cn).to_excel(writer, sheet_name="能源分配测算", index=False)
        batch_df.rename(columns=batch_cn).to_excel(writer, sheet_name="服务批次", index=False)
        customer_summary_df.rename(columns=customer_cn).to_excel(writer, sheet_name="客户需求汇总", index=False)

    wb = load_workbook(result_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_cells in ws.columns:
            max_len = 8
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = cell.value
                if val is None:
                    continue
                max_len = max(max_len, min(len(str(val)) + 2, 36))
                if isinstance(val, float):
                    cell.number_format = "0.0000"
            ws.column_dimensions[col_letter].width = max_len

    # 单独加入图片 sheet
    if "图表" not in wb.sheetnames:
        img_ws = wb.create_sheet("图表")
    else:
        img_ws = wb["图表"]
    positions = ["A1", "A38", "A76"]
    for img_path, anchor in zip(image_paths, positions):
        if img_path.exists():
            img = XLImage(str(img_path))
            img.width = 900
            img.height = int(img.height * (900 / img.width)) if img.width else img.height
            img_ws.add_image(img, anchor)
    wb.save(result_path)


# =========================
# 8. 主程序
# =========================
def solve() -> Dict[str, Any]:
    orders, dist_df, coords_df, tw_df, customer_df = load_input_data(BASE_DIR)
    active_df, batch_df = build_batches(customer_df)
    batch_lookup: Dict[int, Dict[str, Any]] = batch_df.set_index("batch_id").to_dict("index")

    # 初始构造 + 局部搜索 + 内部顺序优化
    routes = build_initial_routes(batch_df, batch_lookup, dist_df)
    routes = local_relocation_search(routes, batch_lookup, dist_df, max_iter=4)
    routes, route_evals, _ = refine_routes_order(routes, batch_lookup, dist_df)

    # 最终线路出发时刻微调
    final_evals = []
    final_routes = []
    for seq in routes:
        best_seq, _ = refine_single_route(seq, batch_lookup, dist_df)
        eva = evaluate_route(best_seq, batch_lookup, dist_df, final_search=True)
        final_routes.append(best_seq)
        final_evals.append(eva)
    routes, route_evals = final_routes, final_evals

    # 新能源车分配
    energy_gain_df, energy_assign_df = select_ev_routes(route_evals, EV_FLEET_LIMIT)

    # 结果表
    route_summary_base, stop_schedule_df, leg_df = build_result_tables(routes, route_evals, energy_assign_df, batch_lookup)
    vehicle_chain_df, vehicle_count = assign_physical_vehicles(route_summary_base)
    route_summary_df = add_vehicle_costs(route_summary_base, vehicle_chain_df)
    vehicle_usage_df = build_vehicle_usage(route_summary_df)

    # 运输流程图和其他图
    plot_transport_flow(route_summary_df, stop_schedule_df, coords_df, FLOW_PNG)
    plot_vehicle_gantt(route_summary_df, GANTT_PNG)
    plot_cost_breakdown(route_summary_df, COST_PNG)

    # 客户汇总表
    customer_summary_df = active_df[[
        "customer_id", "order_count", "weight_kg", "volume_m3", "split_count", "x_km", "y_km", "tw_start_min", "tw_end_min"
    ]].copy()
    customer_summary_df["tw_start"] = customer_summary_df["tw_start_min"].apply(minute_to_hhmm)
    customer_summary_df["tw_end"] = customer_summary_df["tw_end_min"].apply(minute_to_hhmm)

    # 汇总结果
    total_cost = route_summary_df["total_cost"].sum()
    summary_rows = [
        ("订单数", len(orders)),
        ("有需求客户数", len(active_df)),
        ("服务批次数", len(batch_df)),
        ("线路数", len(route_summary_df)),
        ("燃油线路数", int((route_summary_df["vehicle_type"] == "燃油3000kg").sum())),
        ("新能源线路数", int((route_summary_df["vehicle_type"] == "新能源3000kg").sum())),
        ("燃油实体车辆数", vehicle_count.get("燃油3000kg", 0)),
        ("新能源实体车辆数", vehicle_count.get("新能源3000kg", 0)),
        ("实体车辆总数", sum(vehicle_count.values())),
        ("总需求重量kg", active_df["weight_kg"].sum()),
        ("总需求体积m3", active_df["volume_m3"].sum()),
        ("总行驶距离km", route_summary_df["distance_km"].sum()),
        ("总等待时间min", route_summary_df["wait_min"].sum()),
        ("总迟到时间min", route_summary_df["late_min"].sum()),
        ("固定成本元", route_summary_df["fixed_cost"].sum()),
        ("能源成本元", route_summary_df["energy_cost"].sum()),
        ("碳排成本元", route_summary_df["carbon_cost"].sum()),
        ("等待成本元", route_summary_df["wait_cost"].sum()),
        ("迟到成本元", route_summary_df["late_cost"].sum()),
        ("总成本元", total_cost),
        ("燃油车数量约束", f"{vehicle_count.get('燃油3000kg', 0)} / {FUEL_FLEET_LIMIT}"),
        ("新能源车数量约束", f"{vehicle_count.get('新能源3000kg', 0)} / {EV_FLEET_LIMIT}"),
        ("最早出发", minute_to_hhmm(route_summary_df["start_min"].min())),
        ("最晚返回", minute_to_hhmm(route_summary_df["end_min"].max())),
    ]
    summary_df = pd.DataFrame(summary_rows, columns=["指标", "值"])

    export_excel(
        RESULT_XLSX,
        summary_df,
        customer_summary_df,
        batch_df,
        route_summary_df,
        stop_schedule_df,
        leg_df,
        vehicle_chain_df,
        vehicle_usage_df,
        energy_gain_df,
        [FLOW_PNG, GANTT_PNG, COST_PNG],
    )

    summary_text = []
    summary_text.append("问题1求解完成")
    summary_text.append("=" * 72)
    for k, v in summary_rows:
        if isinstance(v, float):
            summary_text.append(f"{k}: {v:.4f}")
        else:
            summary_text.append(f"{k}: {v}")
    summary_text.append("=" * 72)
    summary_text.append(f"结果 Excel: {RESULT_XLSX.name}")
    summary_text.append(f"运输流程图: {FLOW_PNG.name}")
    summary_text.append(f"车辆甘特图: {GANTT_PNG.name}")
    summary_text.append(f"成本结构图: {COST_PNG.name}")
    SUMMARY_TXT.write_text("\n".join(summary_text), encoding="utf-8")
    print("\n".join(summary_text))

    return {
        "summary_df": summary_df,
        "route_summary_df": route_summary_df,
        "stop_schedule_df": stop_schedule_df,
        "vehicle_usage_df": vehicle_usage_df,
        "vehicle_count": vehicle_count,
    }


if __name__ == "__main__":
    solve()
