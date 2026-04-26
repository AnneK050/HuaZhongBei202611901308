# -*- coding: utf-8 -*-
"""
问题2：环保政策影响下的车辆调度策略（异构车队严谨版）

在 q2_solution_route_legal_fuel_labels_only.py 基础上，整合：
  - 完整 5 类异构车队（燃油 3000/1500/1250kg + 新能源 3000/1250kg）
  - 优先级链分配：Must-EV → EV3000(节省) → EV1250(剩余) → 燃油 3000→1500→1250kg(按容量+成本)
  - 按车型容量修正载重率的能耗计算
  - 数量上限硬核查（每类车型独立）
  - 绿色区限行政策（绕行 + 推迟 + 严格核查）

输入文件（与本脚本同目录）：
  订单信息_补全版.xlsx / 距离矩阵.xlsx / 客户坐标信息.xlsx / 时间窗.xlsx
运行：python q2_solution.py
"""

from __future__ import annotations

import heapq
import itertools
import math
import warnings
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# =============================================================================
# 1. 全局参数
# =============================================================================
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR

ORDER_FILE  = "订单信息_补全版.xlsx"
DIST_FILE   = "距离矩阵.xlsx"
COORD_FILE  = "客户坐标信息.xlsx"
TW_FILE     = "时间窗.xlsx"
Q1_RESULT_XLSX = BASE_DIR / "q1_result_fullsplit.xlsx"

DEPOT_ID        = 0
START_OF_DAY    = 8 * 60
END_OF_DAY      = 24 * 60
SERVICE_MIN     = 20
MAX_ROUTE_STOPS = 5

# --- 异构车队规格 ---
# kind: ev/fuel；w_cap: kg；v_cap: m^3；limit: 数量上限；priority: 越小越优先；fixed_cost: 单辆固定启动成本元
FLEET_SPECS: Dict[str, Dict[str, Any]] = {
    "新能源3000kg": {"kind": "ev",   "w_cap": 3000.0, "v_cap": 15.0, "limit": 10, "priority": 1, "fixed_cost": 400.0},
    "新能源1250kg": {"kind": "ev",   "w_cap": 1250.0, "v_cap": 8.5,  "limit": 15, "priority": 2, "fixed_cost": 280.0},
    "燃油3000kg":  {"kind": "fuel", "w_cap": 3000.0, "v_cap": 13.5, "limit": 60, "priority": 3, "fixed_cost": 400.0},
    "燃油1500kg":  {"kind": "fuel", "w_cap": 1500.0, "v_cap": 10.8, "limit": 50, "priority": 4, "fixed_cost": 260.0},
    "燃油1250kg":  {"kind": "fuel", "w_cap": 1250.0, "v_cap": 6.5,  "limit": 50, "priority": 5, "fixed_cost": 220.0},
}

# 通用容量（用于构造阶段的最宽松检查 = 所有车型中最大容量）
MAX_W_CAP = max(spec["w_cap"] for spec in FLEET_SPECS.values())
MAX_V_CAP = max(spec["v_cap"] for spec in FLEET_SPECS.values())
VEHICLE_WEIGHT_CAP = MAX_W_CAP
FUEL_VOLUME_CAP    = max(s["v_cap"] for s in FLEET_SPECS.values() if s["kind"] == "fuel")
EV_VOLUME_CAP      = max(s["v_cap"] for s in FLEET_SPECS.values() if s["kind"] == "ev")
EV_FLEET_LIMIT     = FLEET_SPECS["新能源3000kg"]["limit"]

# --- 能源与碳排 ---
FUEL_PRICE             = 7.61
ELECTRIC_PRICE         = 1.64
CARBON_PRICE           = 0.65
FUEL_CARBON_FACTOR     = 2.547
ELECTRIC_CARBON_FACTOR = 0.501
WAIT_COST_PER_HOUR     = 20.0
LATE_COST_PER_HOUR     = 50.0
CROSS_DAY_PENALTY_PER_MIN = 5000.0

# --- 绿色配送区 ---
GREEN_ZONE_RADIUS  = 10.0
RESTRICT_START_MIN = 8  * 60
RESTRICT_END_MIN   = 16 * 60

TRAFFIC_INTERVALS: List[Tuple[float, float, float]] = [
    (0,              8  * 60, 35.4),
    (8  * 60,        9  * 60,  9.8),
    (9  * 60,        10 * 60, 55.3),
    (10 * 60, int(11.5 * 60), 35.4),
    (int(11.5 * 60), 13 * 60,  9.8),
    (13 * 60,        15 * 60, 55.3),
    (15 * 60,        24 * 60, 35.4),
    (24 * 60,        10 ** 9, 35.4),
]

RESULT_XLSX       = OUTPUT_DIR / "q2_result.xlsx"
FLOW_PNG          = OUTPUT_DIR / "q2_transport_flow_map.png"
GANTT_PNG         = OUTPUT_DIR / "q2_vehicle_gantt.png"
COST_CMP_PNG      = OUTPUT_DIR / "q2_cost_comparison.png"
CARBON_CMP_PNG    = OUTPUT_DIR / "q2_carbon_comparison.png"
SUMMARY_TXT       = OUTPUT_DIR / "q2_console_summary.txt"


# =============================================================================
# 2. 工具函数
# =============================================================================
def hhmm_to_minute(x: Any) -> int:
    if pd.isna(x):
        raise ValueError("时间窗存在空值。")
    if hasattr(x, "hour") and hasattr(x, "minute"):
        return int(x.hour) * 60 + int(x.minute)
    if isinstance(x, (int, float, np.integer, np.floating)):
        v = float(x)
        if 0 <= v <= 1:
            return int(round(v * 24 * 60))
        return int(round(v))
    s = str(x).strip()
    if " " in s:
        s = s.split()[-1]
    parts = s.split(":")
    return int(parts[0]) * 60 + int(parts[1])


def minute_to_hhmm(x: float) -> str:
    x = float(x)
    if not np.isfinite(x):
        return ""
    h = int(x // 60); m = int(round(x % 60))
    if m == 60: h += 1; m = 0
    return f"{h:02d}:{m:02d}"


def fuel_per_100km(v: float) -> float:
    return 0.0025 * v * v - 0.2554 * v + 31.75


def electric_per_100km(v: float) -> float:
    return 0.0014 * v * v - 0.12 * v + 36.19


def set_matplotlib_font() -> None:
    plt.rcParams["font.sans-serif"] = ["DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["figure.dpi"] = 140
    plt.rcParams["savefig.dpi"] = 220


def time_window_cost(eva: Dict[str, Any]) -> float:
    return eva["wait_min"]/60.0*WAIT_COST_PER_HOUR + eva["late_min"]/60.0*LATE_COST_PER_HOUR


def is_ev_type(vt: str) -> bool:
    return FLEET_SPECS.get(str(vt), {}).get("kind") == "ev"


def is_fuel_type(vt: str) -> bool:
    return FLEET_SPECS.get(str(vt), {}).get("kind") == "fuel"


# =============================================================================
# 3. 数据加载与批次拆分
# =============================================================================
def complete_order_weight_volume(orders: pd.DataFrame) -> pd.DataFrame:
    orders = orders.copy()
    orders["weight_kg"] = pd.to_numeric(orders["weight_kg"], errors="coerce")
    orders["volume_m3"] = pd.to_numeric(orders["volume_m3"], errors="coerce")
    vw = orders["weight_kg"].notna() & (orders["weight_kg"] > 0)
    vv = orders["volume_m3"].notna() & (orders["volume_m3"] > 0)
    cm = vw & vv
    lam = float(orders.loc[cm, "volume_m3"].sum() / orders.loc[cm, "weight_kg"].sum())
    mv = vw & ~vv
    orders.loc[mv, "volume_m3"] = orders.loc[mv, "weight_kg"] * lam
    vv2 = orders["volume_m3"].notna() & (orders["volume_m3"] > 0)
    mw = ~vw & vv2
    orders.loc[mw, "weight_kg"] = orders.loc[mw, "volume_m3"] / lam
    orders["weight_kg"] = orders["weight_kg"].fillna(0.0).clip(lower=0.0)
    orders["volume_m3"] = orders["volume_m3"].fillna(0.0).clip(lower=0.0)
    return orders


def split_customer_into_batches(row: pd.Series) -> List[Dict[str, Any]]:
    tw = float(row["weight_kg"]); tv = float(row["volume_m3"])
    if tw <= 1e-9 or tv <= 1e-12:
        return []
    units = max(tw / VEHICLE_WEIGHT_CAP, tv / FUEL_VOLUME_CAP)
    if units <= 1.0 + 1e-9:
        return [{"weight_kg": tw, "volume_m3": tv, "batch_kind": "single",
                 "limiting_dimension": "none", "capacity_fraction": units}]
    lim = "weight" if (tw/VEHICLE_WEIGHT_CAP >= tv/FUEL_VOLUME_CAP) else "volume"
    full = int(math.floor(units + 1e-12))
    uw, uv = tw/units, tv/units
    out = []
    rw, rv = tw, tv
    for _ in range(full):
        bw = min(uw, rw); bv = min(uv, rv)
        if bw <= 1e-9 or bv <= 1e-12: break
        out.append({"weight_kg": bw, "volume_m3": bv, "batch_kind": "full",
                    "limiting_dimension": lim,
                    "capacity_fraction": max(bw/VEHICLE_WEIGHT_CAP, bv/FUEL_VOLUME_CAP)})
        rw -= bw; rv -= bv
    if rw > 1e-7 and rv > 1e-10:
        out.append({"weight_kg": rw, "volume_m3": rv, "batch_kind": "remainder",
                    "limiting_dimension": lim,
                    "capacity_fraction": max(rw/VEHICLE_WEIGHT_CAP, rv/FUEL_VOLUME_CAP)})
    if out:
        dw = tw - sum(b["weight_kg"] for b in out)
        dv = tv - sum(b["volume_m3"] for b in out)
        out[-1]["weight_kg"] += dw; out[-1]["volume_m3"] += dv
        out[-1]["capacity_fraction"] = max(out[-1]["weight_kg"]/VEHICLE_WEIGHT_CAP,
                                            out[-1]["volume_m3"]/FUEL_VOLUME_CAP)
    return out


def load_input_data(base_dir: Path):
    orders = pd.read_excel(base_dir/ORDER_FILE, sheet_name="完整订单数据").rename(
        columns={"目标客户编号": "customer_id", "重量": "weight_kg", "体积": "volume_m3"})
    orders["customer_id"] = orders["customer_id"].astype(int)
    orders = complete_order_weight_volume(orders)

    dr = pd.read_excel(base_dir/DIST_FILE)
    dr = dr.rename(columns={dr.columns[0]: "from_id"})
    dr.columns = ["from_id"] + [int(c) for c in dr.columns[1:]]
    dist_df = dr.set_index("from_id")
    dist_df.index = dist_df.index.astype(int)
    dist_df.columns = dist_df.columns.astype(int)

    coords = pd.read_excel(base_dir/COORD_FILE).rename(
        columns={"ID": "customer_id", "X (km)": "x_km", "Y (km)": "y_km", "类型": "type"})
    coords["customer_id"] = coords["customer_id"].astype(int)

    tw = pd.read_excel(base_dir/TW_FILE).rename(
        columns={"客户编号": "customer_id", "开始时间": "tw_start", "结束时间": "tw_end"})
    tw["customer_id"] = tw["customer_id"].astype(int)
    tw["tw_start_min"] = tw["tw_start"].apply(hhmm_to_minute)
    tw["tw_end_min"]   = tw["tw_end"].apply(hhmm_to_minute)

    agg = orders.groupby("customer_id", as_index=False).agg(
        weight_kg=("weight_kg", "sum"), volume_m3=("volume_m3", "sum"),
        order_count=("customer_id", "count"))
    customer_df = coords.merge(agg, on="customer_id", how="left").merge(
        tw[["customer_id", "tw_start", "tw_end", "tw_start_min", "tw_end_min"]],
        on="customer_id", how="left")
    customer_df["weight_kg"]  = customer_df["weight_kg"].fillna(0.0)
    customer_df["volume_m3"]  = customer_df["volume_m3"].fillna(0.0)
    customer_df["order_count"] = customer_df["order_count"].fillna(0).astype(int)
    return orders, dist_df, coords, tw, customer_df


def build_batches(customer_df: pd.DataFrame):
    active = customer_df[(customer_df["customer_id"] != DEPOT_ID)
                          & (customer_df["weight_kg"] > 1e-9)
                          & (customer_df["volume_m3"] > 1e-12)].copy().sort_values("customer_id").reset_index(drop=True)
    rows = []
    for _, row in active.iterrows():
        pieces = split_customer_into_batches(row)
        sc = len(pieces)
        for k, p in enumerate(pieces, start=1):
            bw, bv = float(p["weight_kg"]), float(p["volume_m3"])
            rows.append({
                "batch_id": len(rows)+1, "customer_id": int(row["customer_id"]),
                "split_index": k, "split_count": sc, "batch_kind": p["batch_kind"],
                "limiting_dimension": p["limiting_dimension"],
                "capacity_fraction": p["capacity_fraction"],
                "weight_kg": bw, "volume_m3": bv,
                "load_ratio_fuel": max(bw/VEHICLE_WEIGHT_CAP, bv/FUEL_VOLUME_CAP),
                "x_km": float(row["x_km"]), "y_km": float(row["y_km"]),
                "tw_start_min": int(row["tw_start_min"]), "tw_end_min": int(row["tw_end_min"]),
                "tw_start": minute_to_hhmm(row["tw_start_min"]),
                "tw_end": minute_to_hhmm(row["tw_end_min"]),
            })
    batch_df = pd.DataFrame(rows)
    active = active.merge(batch_df.groupby("customer_id", as_index=False).agg(
        split_count=("batch_id", "count")), on="customer_id", how="left")
    return active, batch_df


# =============================================================================
# 4. 行驶时间 & 线路评价
# =============================================================================
def travel_time_minutes(distance_km: float, depart: float) -> float:
    if distance_km <= 1e-12: return 0.0
    remain = float(distance_km); t = float(depart); total = 0.0; guard = 0
    while remain > 1e-10:
        guard += 1
        if guard > 1000: break
        for l, r, sp in TRAFFIC_INTERVALS:
            if l <= t < r:
                avail = (r-t)/60.0; possible = sp*avail
                if possible >= remain - 1e-10:
                    cs = remain/sp*60.0
                    total += cs; t += cs; remain = 0.0
                else:
                    total += avail*60.0; remain -= possible; t = r
                break
        else:
            total += remain/TRAFFIC_INTERVALS[-1][2]*60.0; remain = 0.0
    return total


def route_load(seq, batch_lookup):
    return (sum(float(batch_lookup[b]["weight_kg"]) for b in seq),
            sum(float(batch_lookup[b]["volume_m3"]) for b in seq))


def route_capacity_ok(seq, batch_lookup, w_cap=VEHICLE_WEIGHT_CAP, v_cap=FUEL_VOLUME_CAP):
    if len(seq) > MAX_ROUTE_STOPS: return False
    w, v = route_load(seq, batch_lookup)
    return w <= w_cap+1e-9 and v <= v_cap+1e-9


def estimate_route_departure(seq, batch_lookup, dist_df):
    f = batch_lookup[seq[0]]
    d = float(dist_df.loc[DEPOT_ID, int(f["customer_id"])])
    buf = d/35.4*60.0
    return max(float(START_OF_DAY), float(f["tw_start_min"]) - buf)


# 全局坐标缓存（绿色区检测 / 绕行使用）
GLOBAL_COORD_LOOKUP: Dict[int, Dict[str, float]] = {}
LAST_LEG_DF: Optional[pd.DataFrame] = None


def build_green_zone_customers(coords_df: pd.DataFrame) -> Set[int]:
    global GLOBAL_COORD_LOOKUP
    GLOBAL_COORD_LOOKUP = coords_df[["customer_id", "x_km", "y_km"]].set_index("customer_id").to_dict("index")
    return {int(r["customer_id"]) for _, r in coords_df.iterrows()
            if int(r["customer_id"]) != DEPOT_ID
            and math.hypot(float(r["x_km"]), float(r["y_km"])) <= GREEN_ZONE_RADIUS + 1e-9}


def _pt(cid: int) -> np.ndarray:
    r = GLOBAL_COORD_LOOKUP[int(cid)]
    return np.array([float(r["x_km"]), float(r["y_km"])], dtype=float)


def _segment_inside_params(p0, p1, radius=GREEN_ZONE_RADIUS):
    p0 = np.asarray(p0, dtype=float); p1 = np.asarray(p1, dtype=float)
    v = p1 - p0; a = float(v @ v)
    if a <= 1e-12:
        return [(0.0, 1.0)] if np.linalg.norm(p0) <= radius+1e-9 else []
    b = 2.0 * float(p0 @ v); c = float(p0@p0 - radius*radius)
    disc = b*b - 4*a*c; roots = []
    if disc > 1e-10:
        sd = math.sqrt(disc)
        for u in [(-b-sd)/(2*a), (-b+sd)/(2*a)]:
            if -1e-9 <= u <= 1+1e-9:
                roots.append(min(1.0, max(0.0, float(u))))
    elif abs(disc) <= 1e-10:
        u = -b/(2*a)
        if -1e-9 <= u <= 1+1e-9: roots.append(min(1.0, max(0.0, float(u))))
    cuts = sorted(set([0.0, 1.0] + [round(r, 12) for r in roots]))
    out = []
    for a0, a1 in zip(cuts[:-1], cuts[1:]):
        if a1-a0 <= 1e-9: continue
        mid = 0.5*(a0+a1); pm = p0 + mid*v
        if np.linalg.norm(pm) < radius - 1e-8:
            out.append((float(a0), float(a1)))
    if not out and np.linalg.norm(p0) < radius-1e-8 and np.linalg.norm(p1) < radius-1e-8:
        out.append((0.0, 1.0))
    return out


def _interval_overlap(a0, a1, b0, b1):
    return max(float(a0), float(b0)) < min(float(a1), float(b1)) - 1e-7


def _green_windows_for_straight_leg(from_id, to_id, depart_min, raw_dist):
    if not GLOBAL_COORD_LOOKUP: return []
    intervals = _segment_inside_params(_pt(from_id), _pt(to_id))
    res = []
    for u0, u1 in intervals:
        if u1-u0 <= 1e-8: continue
        t0 = float(depart_min) + travel_time_minutes(float(raw_dist)*u0, float(depart_min))
        t1 = float(depart_min) + travel_time_minutes(float(raw_dist)*u1, float(depart_min))
        res.append((t0, t1, u0, u1))
    return res


def _has_restricted_overlap(windows):
    return any(_interval_overlap(w[0], w[1], RESTRICT_START_MIN, RESTRICT_END_MIN) for w in windows)


def _safe_detour_radius(p0, p1):
    dmin = min(float(np.linalg.norm(p0)), float(np.linalg.norm(p1)))
    base = GREEN_ZONE_RADIUS + 0.45
    if dmin > base + 1e-7: return base
    if dmin > GREEN_ZONE_RADIUS + 1e-5:
        return GREEN_ZONE_RADIUS + max(0.01, 0.45*(dmin-GREEN_ZONE_RADIUS))
    return GREEN_ZONE_RADIUS + 0.01


def _angle_delta(a, b, direction):
    a = (a + 2*math.pi) % (2*math.pi); b = (b + 2*math.pi) % (2*math.pi)
    if direction >= 0: return (b-a) % (2*math.pi)
    return (a-b) % (2*math.pi)


def _arc_points(phi0, phi1, direction, radius, n=48):
    delta = _angle_delta(phi0, phi1, direction)
    angles = phi0 + (np.linspace(0, delta, n) if direction >= 0 else -np.linspace(0, delta, n))
    return np.column_stack([radius*np.cos(angles), radius*np.sin(angles)])


def green_detour_path_and_distance(from_id, to_id, raw_dist):
    p0, p1 = _pt(from_id), _pt(to_id)
    eu = float(np.linalg.norm(p1-p0))
    if eu <= 1e-9: return np.vstack([p0, p1]), float(raw_dist)
    radius = _safe_detour_radius(p0, p1)
    d0, d1 = float(np.linalg.norm(p0)), float(np.linalg.norm(p1))
    if d0 <= GREEN_ZONE_RADIUS+1e-8 or d1 <= GREEN_ZONE_RADIUS+1e-8:
        return np.vstack([p0, p1]), float(raw_dist)
    if d0 <= radius+1e-8 or d1 <= radius+1e-8:
        a0 = math.atan2(p0[1], p0[0]); a1 = math.atan2(p1[1], p1[0])
        opts = []
        for direction in [1, -1]:
            arc = _arc_points(a0, a1, direction, radius, n=64)
            geom = float(np.linalg.norm(p0-arc[0]) + np.linalg.norm(p1-arc[-1])
                         + radius*_angle_delta(a0, a1, direction))
            opts.append((geom, np.vstack([p0, arc, p1])))
        geom_dist, path = min(opts, key=lambda x: x[0])
    else:
        th0, th1 = math.atan2(p0[1], p0[0]), math.atan2(p1[1], p1[0])
        g0 = math.acos(radius/d0); g1 = math.acos(radius/d1)
        t0 = math.sqrt(max(0.0, d0*d0 - radius*radius))
        t1 = math.sqrt(max(0.0, d1*d1 - radius*radius))
        best = None
        for phi0 in [th0+g0, th0-g0]:
            for phi1 in [th1+g1, th1-g1]:
                for direction in [1, -1]:
                    arc_d = _angle_delta(phi0, phi1, direction)
                    geom = t0 + t1 + radius*arc_d
                    arc = _arc_points(phi0, phi1, direction, radius, n=72)
                    tp0 = np.array([radius*math.cos(phi0), radius*math.sin(phi0)])
                    tp1 = np.array([radius*math.cos(phi1), radius*math.sin(phi1)])
                    path = np.vstack([p0, tp0, arc, tp1, p1])
                    if best is None or geom < best[0]:
                        best = (geom, path)
        geom_dist, path = best
    scale = max(1.0, float(raw_dist)/max(eu, 1e-9))
    return path, max(float(raw_dist), float(geom_dist)*scale)


def evaluate_route_fixed_start_policy(seq, start_min, batch_lookup, dist_df,
                                       green_zone_customers, use_ev: bool):
    start_min = max(float(START_OF_DAY), float(start_min))
    t = start_min
    lw, lv = route_load(seq, batch_lookup); rw, rv = lw, lv
    td = tw_total = tl = 0.0
    stops, legs = [], []
    prev = DEPOT_ID; viol = False
    for bid in seq:
        row = batch_lookup[bid]; cid = int(row["customer_id"])
        raw_d = float(dist_df.loc[prev, cid]); dep = t
        wins = _green_windows_for_straight_leg(prev, cid, dep, raw_d)
        p0, p1 = _pt(prev), _pt(cid)
        endp_in = (np.linalg.norm(p0) <= GREEN_ZONE_RADIUS+1e-9) or (np.linalg.norm(p1) <= GREEN_ZONE_RADIUS+1e-9)
        rest = (not use_ev) and _has_restricted_overlap(wins)
        detour = bool(rest and (not endp_in))
        if detour:
            _, dist = green_detour_path_and_distance(prev, cid, raw_d)
            drv = travel_time_minutes(dist, dep); arr = dep+drv
            leg_v = False
        else:
            dist = raw_d; drv = travel_time_minutes(dist, dep); arr = dep+drv
            leg_v = bool(rest)
            if leg_v: viol = True
        wait = max(0.0, float(row["tw_start_min"]) - arr)
        ss = arr+wait; late = max(0.0, arr - float(row["tw_end_min"]))
        leave = ss + SERVICE_MIN
        if (not use_ev) and cid in green_zone_customers:
            if _interval_overlap(arr, leave, RESTRICT_START_MIN, RESTRICT_END_MIN):
                viol = True
        td += dist; tw_total += wait; tl += late
        legs.append({"from_id": prev, "to_id": cid, "distance_km": dist,
                     "raw_distance_km": raw_d, "depart_min": dep, "drive_min": drv,
                     "arrival_min": arr, "remain_weight_kg": rw, "remain_volume_m3": rv,
                     "is_detour": detour, "detour_extra_km": max(0.0, dist-raw_d),
                     "straight_crosses_green_zone": bool(wins),
                     "green_restricted_overlap": bool(rest),
                     "leg_policy_violation": leg_v})
        stops.append({"batch_id": bid, "customer_id": cid, "arrival_min": arr,
                      "service_start_min": ss, "leave_min": leave,
                      "wait_min": wait, "late_min": late,
                      "remain_weight_before_kg": rw, "remain_volume_before_m3": rv,
                      "batch_weight_kg": float(row["weight_kg"]),
                      "batch_volume_m3": float(row["volume_m3"])})
        rw -= float(row["weight_kg"]); rv -= float(row["volume_m3"])
        prev = cid; t = leave
    raw_b = float(dist_df.loc[prev, DEPOT_ID]); dep = t
    wins = _green_windows_for_straight_leg(prev, DEPOT_ID, dep, raw_b)
    p0, p1 = _pt(prev), _pt(DEPOT_ID)
    endp_in = (np.linalg.norm(p0) <= GREEN_ZONE_RADIUS+1e-9) or (np.linalg.norm(p1) <= GREEN_ZONE_RADIUS+1e-9)
    rest = (not use_ev) and _has_restricted_overlap(wins)
    detour = bool(rest and (not endp_in))
    if detour:
        _, bd = green_detour_path_and_distance(prev, DEPOT_ID, raw_b)
        bdrv = travel_time_minutes(bd, dep); leg_v = False
    else:
        bd = raw_b; bdrv = travel_time_minutes(bd, dep); leg_v = bool(rest)
        if leg_v: viol = True
    td += bd
    legs.append({"from_id": prev, "to_id": DEPOT_ID, "distance_km": bd,
                 "raw_distance_km": raw_b, "depart_min": dep, "drive_min": bdrv,
                 "arrival_min": dep+bdrv, "remain_weight_kg": max(0, rw),
                 "remain_volume_m3": max(0, rv), "is_detour": detour,
                 "detour_extra_km": max(0, bd-raw_b),
                 "straight_crosses_green_zone": bool(wins),
                 "green_restricted_overlap": bool(rest),
                 "leg_policy_violation": leg_v})
    return {"start_min": start_min, "end_min": dep+bdrv, "distance_km": td,
            "wait_min": tw_total, "late_min": tl, "load_weight_kg": lw,
            "load_volume_m3": lv, "stops": stops, "legs": legs,
            "fuel_policy_violation": viol}


def evaluate_route(seq, batch_lookup, dist_df, start_minute=None, final_search=False, use_ev=True):
    if not seq: return None
    if start_minute is None:
        start_minute = estimate_route_departure(seq, batch_lookup, dist_df)
    eva = evaluate_route_fixed_start_policy(seq, start_minute, batch_lookup, dist_df, set(), use_ev=True)
    # auto shift waits
    for _ in range(20):
        waits = [s["wait_min"] for s in eva["stops"] if s["wait_min"] > 1e-8]
        if not waits: break
        sh = min(waits)
        cand = evaluate_route_fixed_start_policy(seq, eva["start_min"]+sh, batch_lookup, dist_df, set(), use_ev=True)
        if cand["late_min"] <= eva["late_min"]+1e-8 and cand["wait_min"] < eva["wait_min"]-1e-8:
            eva = cand
        else: break
    if final_search:
        center = eva["start_min"]
        best = eva
        best_obj = (time_window_cost(eva) + 0.02*eva["distance_km"]
                    + max(0, eva["end_min"]-END_OF_DAY)*CROSS_DAY_PENALTY_PER_MIN)
        for st in np.arange(max(START_OF_DAY, center-90), min(END_OF_DAY, center+90)+1e-9, 5.0):
            cand = evaluate_route_fixed_start_policy(seq, st, batch_lookup, dist_df, set(), use_ev=True)
            obj = (time_window_cost(cand) + 0.02*cand["distance_km"]
                   + max(0, cand["end_min"]-END_OF_DAY)*CROSS_DAY_PENALTY_PER_MIN)
            if obj < best_obj-1e-9: best, best_obj = cand, obj
        eva = best
    return eva


def evaluate_route_with_green_policy(seq, batch_lookup, dist_df, green_zone_customers, use_ev):
    if use_ev:
        return evaluate_route(seq, batch_lookup, dist_df, final_search=True)
    best = None; best_obj = float("inf")
    fb = None; fb_obj = float("inf")
    starts = list(np.arange(START_OF_DAY, END_OF_DAY+1e-9, 5.0))
    try:
        est = estimate_route_departure(seq, batch_lookup, dist_df)
        starts.extend([est+d for d in np.arange(-30, 35, 5)])
    except Exception: pass
    for st in sorted(set(float(max(START_OF_DAY, min(END_OF_DAY, x))) for x in starts)):
        tr = evaluate_route_fixed_start_policy(seq, st, batch_lookup, dist_df, green_zone_customers, use_ev=False)
        obj = (time_window_cost(tr) + 0.02*tr["distance_km"]
               + max(0, tr["end_min"]-END_OF_DAY)*CROSS_DAY_PENALTY_PER_MIN)
        if obj < fb_obj: fb, fb_obj = tr, obj
        if not tr.get("fuel_policy_violation", False) and obj < best_obj:
            best, best_obj = tr, obj
    if best is None:
        fb["fuel_policy_violation"] = True; return fb
    return best


def penalty_score(eva):
    return eva["distance_km"] + 0.5*eva["wait_min"] + 2.0*eva["late_min"]


# =============================================================================
# 5. 线路构造与局部搜索
# =============================================================================
def build_initial_routes_q2(batch_df, batch_lookup, dist_df, green_zone_customers):
    gz = batch_df[batch_df["customer_id"].isin(green_zone_customers)].sort_values("tw_start_min")["batch_id"].tolist()
    nz = batch_df[~batch_df["customer_id"].isin(green_zone_customers)].sort_values("tw_start_min")["batch_id"].tolist()
    order = gz + nz
    routes: List[List[int]] = []
    for bid in order:
        b = batch_lookup[bid]; is_gz = int(b["customer_id"]) in green_zone_customers
        best = None
        for ridx, seq in enumerate(routes):
            if not route_capacity_ok(seq+[bid], batch_lookup): continue
            seq_has_gz = any(int(batch_lookup[x]["customer_id"]) in green_zone_customers for x in seq)
            if is_gz and not seq_has_gz:
                gz_count = sum(1 for s in routes if any(int(batch_lookup[x]["customer_id"]) in green_zone_customers for x in s))
                if gz_count < EV_FLEET_LIMIT: continue
            be = evaluate_route(seq, batch_lookup, dist_df); bs = penalty_score(be)
            for pos in range(len(seq)+1):
                ns = seq[:pos]+[bid]+seq[pos:]
                ne = evaluate_route(ns, batch_lookup, dist_df, start_minute=be["start_min"])
                sc = (penalty_score(ne)-bs) + 8.0*(ne["late_min"]-be["late_min"])
                if best is None or sc < best[0]: best = (sc, ridx, ns)
        if best is None: routes.append([bid])
        else: routes[best[1]] = best[2]
    return routes


def refine_single_route(seq, batch_lookup, dist_df):
    bs = seq[:]; be = evaluate_route(bs, batch_lookup, dist_df)
    bsc = penalty_score(be) + 5.0*be["late_min"]
    if len(seq) <= 7:
        for perm in itertools.permutations(seq):
            cs = list(perm)
            ce = evaluate_route(cs, batch_lookup, dist_df, start_minute=be["start_min"])
            csc = penalty_score(ce) + 5.0*ce["late_min"]
            if csc < bsc-1e-9: bs, be, bsc = cs, ce, csc
    else:
        imp = True
        while imp:
            imp = False
            for i in range(len(bs)-1):
                for j in range(i+1, len(bs)):
                    cs = bs[:]; cs[i:j+1] = list(reversed(cs[i:j+1]))
                    ce = evaluate_route(cs, batch_lookup, dist_df, start_minute=be["start_min"])
                    csc = penalty_score(ce) + 5.0*ce["late_min"]
                    if csc < bsc-1e-9: bs, be, bsc = cs, ce, csc; imp = True
    return bs, be


def local_relocation_search(routes, batch_lookup, dist_df, max_iter=3):
    imp = True; it = 0
    while imp and it < max_iter:
        imp = False; it += 1
        evs = [evaluate_route(s, batch_lookup, dist_df) for s in routes]
        scs = [penalty_score(e) + 5.0*e["late_min"] for e in evs]
        for i in range(len(routes)):
            for pi, bid in enumerate(routes[i][:]):
                if len(routes[i]) <= 1: continue
                rs = routes[i][:pi]+routes[i][pi+1:]
                re = evaluate_route(rs, batch_lookup, dist_df)
                rsc = penalty_score(re) + 5.0*re["late_min"]
                gained = scs[i]-rsc
                bj=bp=None; bg=1e-6; bsj=None
                for j in range(len(routes)):
                    if j == i: continue
                    if not route_capacity_ok(routes[j]+[bid], batch_lookup): continue
                    bjs = scs[j]
                    for pj in range(len(routes[j])+1):
                        nj = routes[j][:pj]+[bid]+routes[j][pj:]
                        ne = evaluate_route(nj, batch_lookup, dist_df, start_minute=evs[j]["start_min"])
                        nsc = penalty_score(ne) + 5.0*ne["late_min"]
                        ng = gained - (nsc-bjs)
                        if ng > bg: bg, bj, bp, bsj = ng, j, pj, nj
                if bj is not None:
                    routes[i] = rs; routes[bj] = bsj
                    evs[i] = re; scs[i] = rsc
                    nje = evaluate_route(bsj, batch_lookup, dist_df)
                    evs[bj] = nje; scs[bj] = penalty_score(nje) + 5.0*nje["late_min"]
                    imp = True
    return [r for r in routes if r]


# =============================================================================
# 6. 异构车队成本与分配（核心改造）
# =============================================================================
def compute_route_energy_cost_vehicle(eva, vehicle_type):
    spec = FLEET_SPECS[vehicle_type]
    kind = spec["kind"]; w_cap = spec["w_cap"]; v_cap = spec["v_cap"]
    ec = cc = cons = 0.0
    for leg in eva["legs"]:
        d = leg["distance_km"]
        if d <= 1e-12: continue
        drv = leg["drive_min"]
        avg_speed = d/(drv/60.0) if drv > 1e-9 else 35.4
        rw = max(0.0, float(leg["remain_weight_kg"]))
        rv = max(0.0, float(leg["remain_volume_m3"]))
        lr = max(0.0, min(1.0, max(rw/w_cap, rv/v_cap)))
        if kind == "ev":
            kwh = d/100.0 * electric_per_100km(avg_speed) * (1.0 + 0.35*lr)
            ec += kwh*ELECTRIC_PRICE
            cc += kwh*ELECTRIC_CARBON_FACTOR*CARBON_PRICE
            cons += kwh
        else:
            ltr = d/100.0 * fuel_per_100km(avg_speed) * (1.0 + 0.40*lr)
            ec += ltr*FUEL_PRICE
            cc += ltr*FUEL_CARBON_FACTOR*CARBON_PRICE
            cons += ltr
    return float(ec), float(cc), float(cons)


def route_fits_vehicle(eva, vehicle_type):
    spec = FLEET_SPECS[vehicle_type]
    return (float(eva["load_weight_kg"]) <= spec["w_cap"]+1e-9
            and float(eva["load_volume_m3"]) <= spec["v_cap"]+1e-9)


def route_has_restricted_fuel_stop(eva, green):
    for s in eva["stops"]:
        if int(s["customer_id"]) in green and RESTRICT_START_MIN <= float(s["arrival_min"]) < RESTRICT_END_MIN:
            return True
    return False


def route_has_any_green_stop(eva, green):
    for s in eva["stops"]:
        if int(s["customer_id"]) in green: return True
    for leg in eva.get("legs", []):
        if _segment_inside_params(_pt(int(leg["from_id"])), _pt(int(leg["to_id"]))):
            return True
    return False


def _max_concurrent(rows):
    ev = []
    for r in rows:
        ev.append((r["start_min"], 1)); ev.append((r["end_min"], -1))
    ev.sort(key=lambda x: (x[0], x[1]))
    cur = mx = 0
    for _, d in ev:
        cur += d; mx = max(mx, cur)
    return mx


def assign_vehicle_types_q2(routes, route_evals, batch_lookup, dist_df, green_zone_customers):
    """
    严谨异构车队分配（优先级链）：
      1) Must-EV(限行) → 优先 EV3000，超并发降级 EV1250，再不行 → 燃油+推迟/绕行
      2) 可选 EV3000：节省最大且不超 EV3000 并发上限
      3) 可选 EV1250：在剩余容量内继续节省
      4) 燃油：按 3000 → 1500 → 1250kg 优先级，按容量过滤后选**总变量成本+固定成本**最低的可行车型
    """
    n = len(routes)
    must_ev_flags = [route_has_restricted_fuel_stop(eva, green_zone_customers) for eva in route_evals]
    has_green_flags = [route_has_any_green_stop(eva, green_zone_customers) for eva in route_evals]

    # 计算每条线路在每种车型下的变成本（仅用作分配比较）
    var_costs: Dict[Tuple[int, str], float] = {}
    cost_table: List[Dict[str, Any]] = []
    for rid in range(n):
        eva = route_evals[rid]
        row = {"route_id": rid+1}
        for vt in FLEET_SPECS:
            if not route_fits_vehicle(eva, vt):
                row[vt] = float("inf"); var_costs[(rid, vt)] = float("inf")
                continue
            ec, cc, cons = compute_route_energy_cost_vehicle(eva, vt)
            v = ec + cc + time_window_cost(eva)
            var_costs[(rid, vt)] = v
            row[vt+"_var"] = v; row[vt+"_energy"] = ec; row[vt+"_carbon"] = cc; row[vt+"_cons"] = cons
        row["must_ev"] = must_ev_flags[rid]
        row["has_green"] = has_green_flags[rid]
        row["start_min"] = float(eva["start_min"])
        row["end_min"] = float(eva["end_min"])
        cost_table.append(row)

    assignment: Dict[int, str] = {}            # rid -> vehicle_type
    ev_used: Dict[str, List[Dict]] = {"新能源3000kg": [], "新能源1250kg": []}

    def try_assign_ev(rid, vt):
        if var_costs[(rid, vt)] == float("inf"): return False
        cand = ev_used[vt] + [{"start_min": cost_table[rid]["start_min"], "end_min": cost_table[rid]["end_min"]}]
        if _max_concurrent(cand) <= FLEET_SPECS[vt]["limit"]:
            ev_used[vt].append({"start_min": cost_table[rid]["start_min"], "end_min": cost_table[rid]["end_min"]})
            assignment[rid] = vt
            return True
        return False

    # --- 第一档：Must-EV 路线 ---
    must_routes = sorted([i for i in range(n) if must_ev_flags[i]],
                         key=lambda i: -(var_costs[(i, "燃油3000kg")] - var_costs[(i, "新能源3000kg")]))
    must_overflow = []
    for rid in must_routes:
        if try_assign_ev(rid, "新能源3000kg"): continue
        if try_assign_ev(rid, "新能源1250kg"): continue
        must_overflow.append(rid)

    # --- 第二档：可选 EV3000（成本节省最大）---
    optional = [i for i in range(n) if i not in assignment and not must_ev_flags[i]]
    saving3000 = []
    for rid in optional:
        if var_costs[(rid, "燃油3000kg")] == float("inf"): continue
        if var_costs[(rid, "新能源3000kg")] == float("inf"): continue
        s = var_costs[(rid, "燃油3000kg")] - var_costs[(rid, "新能源3000kg")]
        if s > 1e-9: saving3000.append((rid, s))
    saving3000.sort(key=lambda x: -x[1])
    for rid, _ in saving3000:
        try_assign_ev(rid, "新能源3000kg")

    # --- 第三档：可选 EV1250 ---
    optional2 = [i for i in range(n) if i not in assignment and not must_ev_flags[i]]
    saving1250 = []
    for rid in optional2:
        if var_costs[(rid, "新能源1250kg")] == float("inf"): continue
        ref = min(var_costs[(rid, "燃油3000kg")], var_costs[(rid, "燃油1500kg")], var_costs[(rid, "燃油1250kg")])
        if ref == float("inf"): continue
        s = ref - var_costs[(rid, "新能源1250kg")]
        if s > 1e-9: saving1250.append((rid, s))
    saving1250.sort(key=lambda x: -x[1])
    for rid, _ in saving1250:
        try_assign_ev(rid, "新能源1250kg")

    # --- 第四档：燃油优先级回退（含 must_overflow 强制使用燃油）---
    fuel_priority = ["燃油3000kg", "燃油1500kg", "燃油1250kg"]
    for rid in range(n):
        if rid in assignment: continue
        # 选**变量成本+固定成本**最低的可行燃油车型
        best_vt = None; best_total = float("inf")
        for vt in fuel_priority:
            if var_costs[(rid, vt)] == float("inf"): continue
            total = var_costs[(rid, vt)] + FLEET_SPECS[vt]["fixed_cost"]
            if total < best_total - 1e-9:
                best_total = total; best_vt = vt
        if best_vt is None:
            best_vt = "燃油3000kg"  # 兜底
        assignment[rid] = best_vt

    # --- 对分配为燃油且涉及绿色区/穿越的线路重新评价（绕行/推迟）---
    new_evals = []
    policy_violations = 0
    for rid in range(n):
        vt = assignment[rid]; eva = route_evals[rid]
        spec = FLEET_SPECS[vt]
        if spec["kind"] == "fuel" and (must_ev_flags[rid] or has_green_flags[rid] or rid in must_overflow):
            ne = evaluate_route_with_green_policy(routes[rid], batch_lookup, dist_df, green_zone_customers, use_ev=False)
            if ne.get("fuel_policy_violation", False): policy_violations += 1
            new_evals.append(ne)
        elif spec["kind"] == "ev":
            new_evals.append(evaluate_route(routes[rid], batch_lookup, dist_df, final_search=True))
        else:
            new_evals.append(eva)

    # --- 构造分配结果 DataFrame ---
    rows = []
    for rid in range(n):
        vt = assignment[rid]; eva = new_evals[rid]
        if not route_fits_vehicle(eva, vt):
            # 容量校验失败 → 升级到能装下的最大同类车型
            for cand in (["燃油3000kg", "燃油1500kg", "燃油1250kg"] if FLEET_SPECS[vt]["kind"] == "fuel"
                         else ["新能源3000kg", "新能源1250kg"]):
                if route_fits_vehicle(eva, cand):
                    vt = cand; assignment[rid] = vt; break
        ec, cc, cons = compute_route_energy_cost_vehicle(eva, vt)
        # 节省 vs 燃油3000kg 基准
        base_ref = var_costs.get((rid, "燃油3000kg"), float("inf"))
        cur_var = ec + cc + time_window_cost(eva)
        saving = float(base_ref - cur_var) if FLEET_SPECS[vt]["kind"] == "ev" and base_ref != float("inf") else 0.0
        rows.append({
            "route_id": rid+1, "vehicle_type": vt,
            "is_ev": FLEET_SPECS[vt]["kind"] == "ev",
            "must_ev": bool(must_ev_flags[rid]),
            "has_green_stop": bool(has_green_flags[rid]),
            "energy_cost": float(ec), "carbon_cost": float(cc),
            "energy_consumption": float(cons),
            "saving_vs_fuel": float(saving),
        })
    assign_df = pd.DataFrame(rows)
    return assign_df, new_evals, policy_violations


# =============================================================================
# 7. 结果表 / 实体车辆 / 成本
# =============================================================================
def build_result_tables(routes, route_evals, assign_df, batch_lookup, green_zone_customers):
    global LAST_LEG_DF
    rrows, srows, lrows = [], [], []
    for rid, (seq, eva) in enumerate(zip(routes, route_evals), start=1):
        a = assign_df.loc[assign_df["route_id"] == rid].iloc[0]
        vt = a["vehicle_type"]; ec = float(a["energy_cost"]); cc = float(a["carbon_cost"])
        wc = eva["wait_min"]/60.0*WAIT_COST_PER_HOUR
        lc = eva["late_min"]/60.0*LATE_COST_PER_HOUR
        cust_seq = "0-" + "-".join(str(batch_lookup[b]["customer_id"]) for b in seq) + "-0"
        n_green = sum(1 for b in seq if int(batch_lookup[b]["customer_id"]) in green_zone_customers)
        rrows.append({
            "route_id": rid, "vehicle_type": vt, "batch_count": len(seq),
            "green_stops": n_green, "customer_sequence": cust_seq,
            "batch_sequence": "|".join(str(b) for b in seq),
            "load_weight_kg": eva["load_weight_kg"], "load_volume_m3": eva["load_volume_m3"],
            "start_min": eva["start_min"], "end_min": eva["end_min"],
            "start_time": minute_to_hhmm(eva["start_min"]),
            "end_time": minute_to_hhmm(eva["end_min"]),
            "distance_km": eva["distance_km"], "wait_min": eva["wait_min"],
            "late_min": eva["late_min"], "energy_cost": ec, "carbon_cost": cc,
            "energy_consumption": float(a["energy_consumption"]),
            "saving_vs_fuel": float(a["saving_vs_fuel"]),
            "wait_cost": wc, "late_cost": lc,
            "must_ev": bool(a["must_ev"]), "has_green_stop": bool(a["has_green_stop"]),
        })
        for r, s in enumerate(eva["stops"], start=1):
            srows.append({"route_id": rid, "stop_order": r, "batch_id": s["batch_id"],
                          "customer_id": s["customer_id"],
                          "in_green_zone": int(s["customer_id"]) in green_zone_customers,
                          "vehicle_type": vt, "arrival_min": s["arrival_min"],
                          "service_start_min": s["service_start_min"], "leave_min": s["leave_min"],
                          "arrival_time": minute_to_hhmm(s["arrival_min"]),
                          "service_start_time": minute_to_hhmm(s["service_start_min"]),
                          "leave_time": minute_to_hhmm(s["leave_min"]),
                          "wait_min": s["wait_min"], "late_min": s["late_min"],
                          "batch_weight_kg": s["batch_weight_kg"],
                          "batch_volume_m3": s["batch_volume_m3"]})
        for li, leg in enumerate(eva["legs"], start=1):
            avg_sp = (leg["distance_km"]/(leg["drive_min"]/60.0) if leg["drive_min"] > 1e-9 else 35.4)
            lrows.append({"route_id": rid, "leg_order": li, "vehicle_type": vt,
                          "from_id": leg["from_id"], "to_id": leg["to_id"],
                          "distance_km": leg["distance_km"],
                          "raw_distance_km": leg.get("raw_distance_km", leg["distance_km"]),
                          "depart_min": leg["depart_min"], "arrival_min": leg["arrival_min"],
                          "depart_time": minute_to_hhmm(leg["depart_min"]),
                          "arrival_time": minute_to_hhmm(leg["arrival_min"]),
                          "drive_min": leg["drive_min"], "avg_speed_kmh": avg_sp,
                          "remain_weight_kg": leg["remain_weight_kg"],
                          "remain_volume_m3": leg["remain_volume_m3"],
                          "is_detour": bool(leg.get("is_detour", False)),
                          "detour_extra_km": float(leg.get("detour_extra_km", 0.0)),
                          "straight_crosses_green_zone": bool(leg.get("straight_crosses_green_zone", False)),
                          "leg_policy_violation": bool(leg.get("leg_policy_violation", False))})
    rdf = pd.DataFrame(rrows); sdf = pd.DataFrame(srows); ldf = pd.DataFrame(lrows)
    LAST_LEG_DF = ldf.copy()
    return rdf, sdf, ldf


def assign_physical_vehicles(route_summary_df: pd.DataFrame):
    rows = []; counts: Dict[str, int] = {}
    for vt, sub in route_summary_df.groupby("vehicle_type"):
        heap: List[Tuple[float, int]] = []
        nxt = 1; firsts: Dict[int, int] = {}
        for r in sub.sort_values(["start_min", "route_id"]).to_dict("records"):
            if heap and heap[0][0] <= r["start_min"]+1e-9:
                _, vno = heapq.heappop(heap)
            else:
                vno = nxt; firsts[vno] = r["route_id"]; nxt += 1
            heapq.heappush(heap, (r["end_min"], vno))
            rows.append({"route_id": r["route_id"], "vehicle_type": vt,
                         "vehicle_id": f"{vt}-{vno:02d}",
                         "is_new_vehicle": int(firsts.get(vno) == r["route_id"])})
        counts[str(vt)] = nxt-1
    return pd.DataFrame(rows), counts


def add_vehicle_costs(route_summary_df, vehicle_chain_df):
    df = route_summary_df.merge(vehicle_chain_df, on=["route_id", "vehicle_type"], how="left")
    df["fixed_cost"] = df.apply(
        lambda r: FLEET_SPECS[r["vehicle_type"]]["fixed_cost"] if r["is_new_vehicle"] == 1 else 0.0, axis=1)
    df["total_cost"] = df["fixed_cost"] + df["energy_cost"] + df["carbon_cost"] + df["wait_cost"] + df["late_cost"]
    return df


# =============================================================================
# 8. 绘图
# =============================================================================
def plot_greenzone_map(route_summary_df, stop_schedule_df, coords_df, green_zone_customers, out_path):
    set_matplotlib_font()
    coord = coords_df[["customer_id", "x_km", "y_km"]].set_index("customer_id").to_dict("index")
    fig, ax = plt.subplots(figsize=(11, 10.5))
    circle = plt.Circle((0, 0), GREEN_ZONE_RADIUS, fill=True, facecolor="#e8f5e9",
                         edgecolor="#2ca02c", linewidth=1.6, linestyle="--", alpha=0.32,
                         label=f"Green Zone (r={GREEN_ZONE_RADIUS}km)")
    ax.add_patch(circle)
    non_gz = coords_df[(coords_df["customer_id"] != DEPOT_ID)
                       & (~coords_df["customer_id"].isin(green_zone_customers))]
    gz = coords_df[coords_df["customer_id"].isin(green_zone_customers)]
    ax.scatter(non_gz["x_km"], non_gz["y_km"], s=18, color="#7f7f7f", alpha=0.55, label="Non-GZ")
    ax.scatter(gz["x_km"], gz["y_km"], s=42, color="#2ca02c", alpha=0.9, label="GZ")
    ax.scatter(coord[DEPOT_ID]["x_km"], coord[DEPOT_ID]["y_km"], marker="*", s=320, color="red", zorder=8, label="Depot")
    color_map = {"燃油3000kg": "tab:orange", "燃油1500kg": "tab:red", "燃油1250kg": "tab:purple",
                 "新能源3000kg": "tab:blue", "新能源1250kg": "tab:cyan"}
    type_lu = route_summary_df.set_index("route_id")["vehicle_type"].to_dict()
    done = set()
    leg_df = LAST_LEG_DF
    if leg_df is not None and not leg_df.empty:
        for _, row in leg_df.sort_values(["route_id", "leg_order"]).iterrows():
            vt = row["vehicle_type"]; color = color_map.get(vt, "gray")
            from_id, to_id = int(row["from_id"]), int(row["to_id"])
            p0 = np.array([coord[from_id]["x_km"], coord[from_id]["y_km"]])
            p1 = np.array([coord[to_id]["x_km"], coord[to_id]["y_km"]])
            lab = vt if vt not in done else None; done.add(vt)
            if bool(row.get("is_detour", False)):
                path, _ = green_detour_path_and_distance(from_id, to_id, float(row.get("raw_distance_km", row["distance_km"])))
                ax.plot(path[:, 0], path[:, 1], color=color, lw=1.3, alpha=0.6, label=lab)
            else:
                ax.plot([p0[0], p1[0]], [p0[1], p1[1]], color=color, lw=0.7, alpha=0.28, label=lab)
    ax.scatter(0, 0, marker="+", s=120, color="black", label="City Center")
    ax.set_aspect("equal", adjustable="box")
    ax.set_title("Q2 Routes with Green Zone Policy (Heterogeneous Fleet)")
    ax.set_xlabel("X (km)"); ax.set_ylabel("Y (km)")
    ax.grid(alpha=0.25, linestyle="--"); ax.legend(loc="upper right", fontsize=7)
    plt.tight_layout(); fig.savefig(out_path, bbox_inches="tight"); plt.close(fig)


def plot_vehicle_gantt(route_summary_df, out_path):
    set_matplotlib_font()
    df = route_summary_df.sort_values(["vehicle_type", "vehicle_id", "start_min"]).copy()
    df["vl"] = df["vehicle_id"]
    vorder = df["vl"].drop_duplicates().tolist()
    y_map = {v: i for i, v in enumerate(vorder)}
    cmap = {"燃油3000kg": "tab:orange", "燃油1500kg": "tab:red", "燃油1250kg": "tab:purple",
            "新能源3000kg": "tab:blue", "新能源1250kg": "tab:cyan"}
    fig, ax = plt.subplots(figsize=(14, max(7, 0.22*len(vorder))))
    for _, r in df.iterrows():
        y = y_map[r["vl"]]
        ax.barh(y=y, width=r["end_min"]-r["start_min"], left=r["start_min"],
                height=0.55, color=cmap.get(r["vehicle_type"], "gray"), alpha=0.78)
        ax.text(r["start_min"], y, f"R{int(r['route_id'])}", va="center", ha="left", fontsize=6)
    ax.axvspan(RESTRICT_START_MIN, RESTRICT_END_MIN, alpha=0.08, color="green", label="Fuel Ban 8-16h")
    xt = list(range(8*60, 23*60+1, 60))
    ax.set_xticks(xt); ax.set_xticklabels([minute_to_hhmm(x) for x in xt], rotation=45)
    ax.set_yticks(list(y_map.values())); ax.set_yticklabels(vorder, fontsize=6)
    ax.set_title("Q2 Vehicle Gantt (Heterogeneous Fleet)")
    ax.set_xlabel("Time"); ax.grid(axis="x", alpha=0.25, linestyle="--"); ax.legend(fontsize=8)
    plt.tight_layout(); fig.savefig(out_path, bbox_inches="tight"); plt.close(fig)


def plot_cost_comparison(q1m, q2m, out_path):
    set_matplotlib_font()
    keys = ["Fixed", "Energy", "Carbon", "Waiting", "Late", "Total"]
    q1v = [q1m.get(k, 0) for k in keys]; q2v = [q2m.get(k, 0) for k in keys]
    x = np.arange(len(keys)); w = 0.35
    fig, ax = plt.subplots(figsize=(11, 6))
    ax.bar(x-w/2, q1v, w, label="Q1", color="tab:blue", alpha=0.8)
    ax.bar(x+w/2, q2v, w, label="Q2", color="tab:orange", alpha=0.8)
    ax.set_xticks(x); ax.set_xticklabels(keys); ax.legend(); ax.grid(axis="y", alpha=0.25, linestyle="--")
    ax.set_title("Q1 vs Q2 Cost Comparison")
    plt.tight_layout(); fig.savefig(out_path, bbox_inches="tight"); plt.close(fig)


def plot_carbon_comparison(q1m, q2m, out_path):
    set_matplotlib_font()
    cats = ["Fuel CO2", "EV CO2", "Total CO2"]
    q1v = [q1m.get("fuel_carbon_kg", 0), q1m.get("ev_carbon_kg", 0), q1m.get("total_carbon_kg", 0)]
    q2v = [q2m.get("fuel_carbon_kg", 0), q2m.get("ev_carbon_kg", 0), q2m.get("total_carbon_kg", 0)]
    x = np.arange(len(cats)); w = 0.35
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(x-w/2, q1v, w, label="Q1", color="tab:blue", alpha=0.8)
    ax.bar(x+w/2, q2v, w, label="Q2", color="tab:orange", alpha=0.8)
    ax.set_xticks(x); ax.set_xticklabels(cats); ax.legend(); ax.grid(axis="y", alpha=0.25, linestyle="--")
    ax.set_title("Q1 vs Q2 Carbon Comparison")
    plt.tight_layout(); fig.savefig(out_path, bbox_inches="tight"); plt.close(fig)


# =============================================================================
# 9. Excel 导出
# =============================================================================
def export_excel_q2(path, summary_df, comparison_df, route_summary_df, stop_schedule_df,
                     leg_df, vehicle_chain_df, vehicle_usage_df, image_paths):
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        summary_df.to_excel(w, sheet_name="汇总结果", index=False)
        comparison_df.to_excel(w, sheet_name="问题1vs2对比", index=False)
        route_summary_df.to_excel(w, sheet_name="线路汇总", index=False)
        stop_schedule_df.to_excel(w, sheet_name="停靠时刻表", index=False)
        leg_df.to_excel(w, sheet_name="路段明细", index=False)
        vehicle_chain_df.to_excel(w, sheet_name="线路车辆映射", index=False)
        vehicle_usage_df.to_excel(w, sheet_name="实体车辆使用", index=False)

    wb = load_workbook(path)
    hf = PatternFill("solid", fgColor="1F4E78"); hfont = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for c in ws[1]:
            c.fill = hf; c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            ml = 8; cl = get_column_letter(col[0].column)
            for c in col:
                if c.value is None: continue
                ml = max(ml, min(len(str(c.value))+2, 36))
                if isinstance(c.value, float): c.number_format = "0.0000"
            ws.column_dimensions[cl].width = ml
    if "图表" not in wb.sheetnames:
        iws = wb.create_sheet("图表")
    else:
        iws = wb["图表"]
    pos = ["A1", "A40", "A80", "A120"]
    for ip, an in zip(image_paths, pos):
        if ip.exists():
            img = XLImage(str(ip))
            if img.width and img.width > 0:
                img.height = int(img.height * (900/img.width)); img.width = 900
            iws.add_image(img, an)
    wb.save(path)


def load_q1_metrics(p):
    if not p.exists(): return None
    try:
        s = pd.read_excel(p, sheet_name="汇总结果")
        m = {}
        for _, r in s.iterrows():
            try: m[str(r.iloc[0])] = float(r.iloc[1])
            except: m[str(r.iloc[0])] = 0.0
        rd = pd.read_excel(p, sheet_name="线路汇总")
        def col(kw, default):
            return next((c for c in rd.columns if kw in str(c)), default)
        ec = col("能源成本", "energy_cost"); cc = col("碳排成本", "carbon_cost")
        fc = col("固定成本", "fixed_cost"); tc = col("总成本", "total_cost")
        wc = col("等待成本", "wait_cost"); lc = col("迟到成本", "late_cost")
        tp = col("车辆类型", "vehicle_type"); dc = col("距离", "distance_km")
        return {"Fixed": float(rd[fc].sum()), "Energy": float(rd[ec].sum()),
                "Carbon": float(rd[cc].sum()), "Waiting": float(rd[wc].sum()),
                "Late": float(rd[lc].sum()), "Total": float(rd[tc].sum()),
                "total_distance_km": float(rd[dc].sum()),
                "total_carbon_kg": float(rd[cc].sum())/CARBON_PRICE,
                "fuel_carbon_kg": float(rd.loc[rd[tp].astype(str).str.contains("燃油"), cc].sum())/CARBON_PRICE,
                "ev_carbon_kg": float(rd.loc[rd[tp].astype(str).str.contains("新能源"), cc].sum())/CARBON_PRICE,
                "fuel_routes": int(rd[tp].astype(str).str.contains("燃油").sum()),
                "ev_routes": int(rd[tp].astype(str).str.contains("新能源").sum())}
    except Exception as e:
        print(f"[警告] 读取Q1结果失败: {e}"); return None


# =============================================================================
# 10. 主程序
# =============================================================================
def solve():
    print("="*72); print("问题2：异构车队严谨版"); print("="*72)
    print("[1/8] 加载数据...")
    orders, dist_df, coords_df, _, customer_df = load_input_data(BASE_DIR)
    active_df, batch_df = build_batches(customer_df)
    batch_lookup = batch_df.set_index("batch_id").to_dict("index")
    green = build_green_zone_customers(coords_df)
    print(f"    绿色区客户: {len(green)} / 总批次: {len(batch_df)}")

    print("[2/8] 构造初始线路...")
    routes = build_initial_routes_q2(batch_df, batch_lookup, dist_df, green)

    print("[3/8] 局部重定位搜索...")
    routes = local_relocation_search(routes, batch_lookup, dist_df, max_iter=3)

    print("[4/8] 单线路优化...")
    refined_routes, refined_evals = [], []
    for seq in routes:
        bs, _ = refine_single_route(seq, batch_lookup, dist_df)
        refined_routes.append(bs)
        refined_evals.append(evaluate_route(bs, batch_lookup, dist_df, final_search=True))
    routes, route_evals = refined_routes, refined_evals

    print("[5/8] 异构车队车型分配（5档优先级）...")
    assign_df, route_evals, viol = assign_vehicle_types_q2(routes, route_evals, batch_lookup, dist_df, green)
    if viol > 0:
        print(f"    [警告] {viol} 条燃油车线路因时间窗过紧标记为政策违规")
    for vt in FLEET_SPECS:
        n = int((assign_df["vehicle_type"] == vt).sum())
        print(f"    {vt}: {n} 条线路")

    print("[6/8] 构建结果表 + 实体车辆指派...")
    rsum_df, stop_df, leg_df = build_result_tables(routes, route_evals, assign_df, batch_lookup, green)
    chain_df, vcount = assign_physical_vehicles(rsum_df)

    # 数量上限硬核查
    for vt, spec in FLEET_SPECS.items():
        n = vcount.get(vt, 0)
        if n > spec["limit"]:
            raise RuntimeError(f"{vt} 实体车辆数 {n} 超过上限 {spec['limit']}")

    rsum_df = add_vehicle_costs(rsum_df, chain_df)
    vu = []
    for vid, sub in rsum_df.groupby("vehicle_id"):
        vu.append({"vehicle_id": vid, "vehicle_type": sub["vehicle_type"].iloc[0],
                   "route_count": len(sub),
                   "first_start_time": minute_to_hhmm(sub["start_min"].min()),
                   "last_end_time": minute_to_hhmm(sub["end_min"].max()),
                   "distance_km": sub["distance_km"].sum(),
                   "total_cost": sub["total_cost"].sum()})
    vu_df = pd.DataFrame(vu)

    print("[7/8] 生成图表...")
    plot_greenzone_map(rsum_df, stop_df, coords_df, green, FLOW_PNG)
    plot_vehicle_gantt(rsum_df, GANTT_PNG)
    q1m = load_q1_metrics(Q1_RESULT_XLSX)
    q2_fc = float(rsum_df.loc[rsum_df["vehicle_type"].astype(str).str.contains("燃油"), "carbon_cost"].sum())/CARBON_PRICE
    q2_ec = float(rsum_df.loc[rsum_df["vehicle_type"].astype(str).str.contains("新能源"), "carbon_cost"].sum())/CARBON_PRICE
    q2m = {"Fixed": float(rsum_df["fixed_cost"].sum()),
           "Energy": float(rsum_df["energy_cost"].sum()),
           "Carbon": float(rsum_df["carbon_cost"].sum()),
           "Waiting": float(rsum_df["wait_cost"].sum()),
           "Late": float(rsum_df["late_cost"].sum()),
           "Total": float(rsum_df["total_cost"].sum()),
           "total_distance_km": float(rsum_df["distance_km"].sum()),
           "total_carbon_kg": q2_fc + q2_ec,
           "fuel_carbon_kg": q2_fc, "ev_carbon_kg": q2_ec,
           "fuel_routes": int(rsum_df["vehicle_type"].astype(str).str.contains("燃油").sum()),
           "ev_routes": int(rsum_df["vehicle_type"].astype(str).str.contains("新能源").sum())}
    if q1m:
        plot_cost_comparison(q1m, q2m, COST_CMP_PNG)
        plot_carbon_comparison(q1m, q2m, CARBON_CMP_PNG)
    else:
        zero = {k: 0 for k in ["Fixed", "Energy", "Carbon", "Waiting", "Late", "Total",
                                "fuel_carbon_kg", "ev_carbon_kg", "total_carbon_kg"]}
        plot_cost_comparison(zero, q2m, COST_CMP_PNG)
        plot_carbon_comparison(zero, q2m, CARBON_CMP_PNG)

    print("[8/8] 导出 Excel + 汇总文本...")
    summary_rows = [
        ("有需求客户数", len(active_df)),
        ("服务批次数", len(batch_df)),
        ("绿色区客户数", len(green)),
        ("线路总数", len(rsum_df)),
    ]
    for vt in sorted(FLEET_SPECS, key=lambda x: FLEET_SPECS[x]["priority"]):
        n_route = int((rsum_df["vehicle_type"] == vt).sum())
        n_veh = vcount.get(vt, 0)
        summary_rows.append((f"{vt}线路数", n_route))
        summary_rows.append((f"{vt}实体车辆", f"{n_veh}/{FLEET_SPECS[vt]['limit']}"))
    summary_rows.extend([
        ("实体车辆总数", sum(vcount.values())),
        ("总行驶距离km", round(q2m["total_distance_km"], 4)),
        ("固定成本元", round(q2m["Fixed"], 2)),
        ("能源成本元", round(q2m["Energy"], 2)),
        ("碳排成本元", round(q2m["Carbon"], 2)),
        ("等待成本元", round(q2m["Waiting"], 2)),
        ("迟到成本元", round(q2m["Late"], 2)),
        ("总成本元", round(q2m["Total"], 2)),
        ("总碳排放kg", round(q2m["total_carbon_kg"], 2)),
        ("政策违规线路数", viol),
    ])
    summary_df = pd.DataFrame(summary_rows, columns=["指标", "值"])

    cmp_rows = []
    if q1m:
        def d(a, b):
            return f"{(a-b)/abs(b)*100:+.2f}%" if abs(b) > 1e-9 else "N/A"
        for k, lab in [("Total", "总成本元"), ("Fixed", "固定成本元"), ("Energy", "能源成本元"),
                        ("Carbon", "碳排成本元"), ("Waiting", "等待成本元"), ("Late", "迟到成本元")]:
            cmp_rows.append({"指标": lab, "问题1": q1m.get(k, 0), "问题2": q2m[k],
                              "变化": d(q2m[k], q1m.get(k, 0))})
        cmp_rows.append({"指标": "总碳排放kg", "问题1": round(q1m.get("total_carbon_kg", 0), 2),
                          "问题2": round(q2m["total_carbon_kg"], 2),
                          "变化": d(q2m["total_carbon_kg"], q1m.get("total_carbon_kg", 0))})
    cmp_df = pd.DataFrame(cmp_rows) if cmp_rows else pd.DataFrame({"说明": ["未找到Q1结果"]})

    export_excel_q2(RESULT_XLSX, summary_df, cmp_df, rsum_df, stop_df, leg_df,
                     chain_df, vu_df, [FLOW_PNG, GANTT_PNG, COST_CMP_PNG, CARBON_CMP_PNG])

    lines = ["="*72, "问题2 异构车队严谨版 — 结果汇总", "="*72]
    for k, v in summary_rows:
        lines.append(f"  {k}: {v:.4f}" if isinstance(v, float) else f"  {k}: {v}")
    if cmp_rows:
        lines.append(""); lines.append("【Q1 vs Q2 对比】")
        for r in cmp_rows:
            lines.append(f"  {r['指标']:14s} Q1={r['问题1']}  Q2={r['问题2']}  Δ={r['变化']}")
    lines.append("="*72)
    SUMMARY_TXT.write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(lines))
    return {"route_summary_df": rsum_df, "vehicle_count": vcount, "q2_metrics": q2m}


if __name__ == "__main__":
    solve()